# ─────────────────────────────────────────────────────────────
#  ui/exportar.py  –  Panel de exportación a Excel
# ─────────────────────────────────────────────────────────────

import os
import sys
import shutil
from datetime import date, datetime

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QFileDialog, QMessageBox, QScrollArea,
    QSizePolicy, QSpinBox, QLineEdit,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db import database as db
from ui.styles import _SPIN_SUBCONTROLES as _SPIN_SUBS


# ─────────────────────────────────────────────────────────────
#  Worker para no bloquear la UI mientras genera el Excel
# ─────────────────────────────────────────────────────────────

class ExportWorker(QThread):
    terminado = pyqtSignal(bool, str)   # (exito, ruta_o_error)

    def __init__(self, tipo: str, ruta: str, periodo: str = "historico"):
        super().__init__()
        self.tipo    = tipo
        self.ruta    = ruta
        self.periodo = periodo

    def run(self):
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # quitar hoja vacía por defecto

            # Colores tema vinoteca
            COLOR_ENCABEZADO  = "722F37"   # burdeos
            COLOR_TOTAL       = "4A1520"   # burdeos oscuro
            COLOR_FILA_PAR    = "F5F0F1"   # rosa muy suave
            COLOR_TEXTO_BLANC = "FFFFFF"

            def _estilo_encabezado(cell):
                cell.font      = Font(bold=True, color=COLOR_TEXTO_BLANC, size=10)
                cell.fill      = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = Border(
                    bottom=Side(style="thin", color="FFFFFF"),
                    right=Side(style="thin",  color="FFFFFF"),
                )

            def _estilo_total(cell):
                cell.font      = Font(bold=True, color=COLOR_TEXTO_BLANC, size=10)
                cell.fill      = PatternFill("solid", fgColor=COLOR_TOTAL)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            def _auto_ancho(ws, min_w=8, max_w=50):
                for col_cells in ws.columns:
                    largo = max(
                        len(str(c.value)) if c.value is not None else 0
                        for c in col_cells
                    )
                    ws.column_dimensions[
                        get_column_letter(col_cells[0].column)
                    ].width = min(max(largo + 3, min_w), max_w)

            def _freeze(ws, cell="A2"):
                ws.freeze_panes = cell

            if self.tipo == "ventas_resumen":
                self._export_ventas_resumen(wb, _estilo_encabezado, _estilo_total,
                                            _auto_ancho, _freeze, COLOR_FILA_PAR)
            elif self.tipo == "ventas_detalle":
                self._export_ventas_detalle(wb, _estilo_encabezado, _estilo_total,
                                            _auto_ancho, _freeze, COLOR_FILA_PAR)
            elif self.tipo == "stock":
                self._export_stock(wb, _estilo_encabezado, _estilo_total,
                                   _auto_ancho, _freeze, COLOR_FILA_PAR)
            elif self.tipo == "proveedores":
                self._export_proveedores(wb, _estilo_encabezado, _estilo_total,
                                         _auto_ancho, _freeze, COLOR_FILA_PAR)
            elif self.tipo == "completo":
                self._export_ventas_resumen(wb, _estilo_encabezado, _estilo_total,
                                            _auto_ancho, _freeze, COLOR_FILA_PAR)
                self._export_ventas_detalle(wb, _estilo_encabezado, _estilo_total,
                                            _auto_ancho, _freeze, COLOR_FILA_PAR)
                self._export_stock(wb, _estilo_encabezado, _estilo_total,
                                   _auto_ancho, _freeze, COLOR_FILA_PAR)
                self._export_proveedores(wb, _estilo_encabezado, _estilo_total,
                                         _auto_ancho, _freeze, COLOR_FILA_PAR)

            wb.save(self.ruta)
            self.terminado.emit(True, self.ruta)

        except Exception as e:
            self.terminado.emit(False, str(e))

    # ── Helpers de período ─────────────────────────────────────

    def _rango_fechas(self):
        """Devuelve (desde, hasta) como strings 'YYYY-MM-DD' según self.periodo."""
        hoy = date.today()
        if self.periodo == "hoy":
            return str(hoy), str(hoy)
        elif self.periodo == "semana":
            from datetime import timedelta
            lunes = hoy - __import__("datetime").timedelta(days=hoy.weekday())
            return str(lunes), str(hoy)
        elif self.periodo == "mes":
            return f"{hoy.year}-{hoy.month:02d}-01", str(hoy)
        elif self.periodo == "anio":
            return f"{hoy.year}-01-01", str(hoy)
        else:  # historico
            return "2000-01-01", str(hoy)

    # ── Hoja: Ventas resumen por día ───────────────────────────

    def _export_ventas_resumen(self, wb, enc, tot, ancho, freeze, fila_par):
        from openpyxl.styles import PatternFill, Alignment, Font, numbers
        ws = wb.create_sheet("Ventas por día")
        freeze(ws)

        desde, hasta = self._rango_fechas()
        conn = db.get_connection()
        rows = conn.execute("""
            SELECT
                v.fecha,
                strftime('%d/%m/%Y', v.fecha)      AS fecha_fmt,
                COUNT(*)                            AS cant_ventas,
                SUM(v.total)                        AS total_dia,
                SUM(v.descuento)                    AS descuentos,
                SUM(CASE WHEN v.medio_pago='efectivo'      THEN v.total ELSE 0 END) AS efectivo,
                SUM(CASE WHEN v.medio_pago='debito'        THEN v.total ELSE 0 END) AS debito,
                SUM(CASE WHEN v.medio_pago='credito'       THEN v.total ELSE 0 END) AS credito,
                SUM(CASE WHEN v.medio_pago='transferencia' THEN v.total ELSE 0 END) AS transferencia,
                SUM(CASE WHEN v.medio_pago='qr'            THEN v.total ELSE 0 END) AS qr
            FROM ventas v
            WHERE v.anulada = 0 AND v.fecha BETWEEN ? AND ?
            GROUP BY v.fecha
            ORDER BY v.fecha
        """, (desde, hasta)).fetchall()
        conn.close()

        encabezados = ["Fecha", "N° Ventas", "Total del día", "Descuentos",
                       "Efectivo", "Débito", "Crédito", "Transferencia", "QR"]
        for col, h in enumerate(encabezados, 1):
            c = ws.cell(row=1, column=col, value=h)
            enc(c)
        ws.row_dimensions[1].height = 28

        total_acum = [0.0] * (len(encabezados) - 2)   # cols numéricas
        for r, row in enumerate(rows, 2):
            valores = [
                row["fecha_fmt"],
                row["cant_ventas"],
                row["total_dia"]       or 0,
                row["descuentos"]      or 0,
                row["efectivo"]        or 0,
                row["debito"]          or 0,
                row["credito"]         or 0,
                row["transferencia"]   or 0,
                row["qr"]              or 0,
            ]
            for i, n in enumerate(valores[2:]):
                total_acum[i] += n

            fill = PatternFill("solid", fgColor=fila_par) if r % 2 == 0 else None
            for col, val in enumerate(valores, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = Alignment(horizontal="right" if col > 1 else "left",
                                        vertical="center")
                if col > 2:
                    c.number_format = '#,##0.00'
                if fill:
                    c.fill = fill

        # Fila TOTAL
        fila_tot = len(rows) + 2
        ws.cell(row=fila_tot, column=1, value="TOTAL")
        ws.cell(row=fila_tot, column=2, value=sum(r["cant_ventas"] for r in rows))
        for i, v in enumerate(total_acum):
            c = ws.cell(row=fila_tot, column=i + 3, value=v)
            c.number_format = '#,##0.00'
        for col in range(1, len(encabezados) + 1):
            tot(ws.cell(row=fila_tot, column=col))

        ancho(ws)

    # ── Hoja: Ventas detalle (ítem por ítem) ───────────────────

    def _export_ventas_detalle(self, wb, enc, tot, ancho, freeze, fila_par):
        from openpyxl.styles import PatternFill, Alignment
        ws = wb.create_sheet("Ventas detalle")
        freeze(ws)

        desde, hasta = self._rango_fechas()
        conn = db.get_connection()
        rows = conn.execute("""
            SELECT
                strftime('%d/%m/%Y', v.fecha) AS fecha,
                v.hora,
                v.id                           AS n_venta,
                v.medio_pago,
                p.nombre                       AS producto,
                c.nombre                       AS categoria,
                dv.cantidad,
                dv.precio_unit,
                dv.subtotal,
                dv.descuento,
                v.total                        AS total_venta,
                v.notas
            FROM ventas v
            JOIN detalle_ventas dv ON dv.venta_id = v.id
            JOIN productos p       ON p.id = dv.producto_id
            LEFT JOIN categorias c ON c.id = p.categoria_id
            WHERE v.anulada = 0 AND v.fecha BETWEEN ? AND ?
            ORDER BY v.fecha, v.id, p.nombre
        """, (desde, hasta)).fetchall()
        conn.close()

        encabezados = ["Fecha", "Hora", "N° Venta", "Medio pago",
                       "Producto", "Categoría", "Cantidad",
                       "Precio unit.", "Subtotal ítem", "Descuento ítem",
                       "Total venta", "Notas"]
        for col, h in enumerate(encabezados, 1):
            c = ws.cell(row=1, column=col, value=h)
            enc(c)
        ws.row_dimensions[1].height = 28

        nums = {7, 8, 9, 10, 11}
        for r, row in enumerate(rows, 2):
            vals = [
                row["fecha"], row["hora"], row["n_venta"], row["medio_pago"],
                row["producto"], row["categoria"], row["cantidad"],
                row["precio_unit"], row["subtotal"], row["descuento"],
                row["total_venta"], row["notas"] or "",
            ]
            fill = PatternFill("solid", fgColor=fila_par) if r % 2 == 0 else None
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = Alignment(horizontal="right" if col in nums else "left",
                                        vertical="center")
                if col in nums:
                    c.number_format = '#,##0.00'
                if fill:
                    c.fill = fill

        ancho(ws)

    # ── Hoja: Stock actual ─────────────────────────────────────

    def _export_stock(self, wb, enc, tot, ancho, freeze, fila_par):
        from openpyxl.styles import PatternFill, Alignment, Font
        ws = wb.create_sheet("Stock actual")
        freeze(ws)

        conn = db.get_connection()
        rows = conn.execute("""
            SELECT
                p.nombre,
                c.nombre           AS categoria,
                pr.nombre          AS proveedor,
                p.codigo_barras,
                p.precio_venta,
                p.precio_costo,
                p.stock_actual,
                p.stock_minimo,
                p.unidad,
                CASE WHEN p.stock_actual <= p.stock_minimo THEN 'BAJO' ELSE 'OK' END AS alerta
            FROM productos p
            LEFT JOIN categorias c  ON c.id  = p.categoria_id
            LEFT JOIN proveedores pr ON pr.id = p.proveedor_id
            WHERE p.activo = 1
            ORDER BY c.nombre, p.nombre
        """).fetchall()
        conn.close()

        encabezados = ["Producto", "Categoría", "Proveedor", "Cód. barras",
                       "Precio venta", "Precio costo", "Stock actual",
                       "Stock mínimo", "Unidad", "Alerta"]
        for col, h in enumerate(encabezados, 1):
            c = ws.cell(row=1, column=col, value=h)
            enc(c)
        ws.row_dimensions[1].height = 28

        nums = {5, 6, 7, 8}
        for r, row in enumerate(rows, 2):
            vals = [
                row["nombre"], row["categoria"] or "", row["proveedor"] or "",
                row["codigo_barras"] or "", row["precio_venta"], row["precio_costo"],
                row["stock_actual"], row["stock_minimo"], row["unidad"],
                row["alerta"],
            ]
            alerta_bajo = row["alerta"] == "BAJO"
            fill = PatternFill("solid", fgColor="FFD7D7") if alerta_bajo \
                   else PatternFill("solid", fgColor=fila_par) if r % 2 == 0 \
                   else None
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = Alignment(horizontal="right" if col in nums else "left",
                                        vertical="center")
                if col in {5, 6}:
                    c.number_format = '#,##0.00'
                if fill:
                    c.fill = fill
                if col == 10 and alerta_bajo:
                    c.font = Font(bold=True, color="CC0000")

        ancho(ws)

    # ── Hoja: Cuentas proveedores ──────────────────────────────

    def _export_proveedores(self, wb, enc, tot, ancho, freeze, fila_par):
        from openpyxl.styles import PatternFill, Alignment, Font
        ws = wb.create_sheet("Cuentas proveedores")
        freeze(ws)

        conn = db.get_connection()
        rows = conn.execute("""
            SELECT
                pr.nombre                              AS proveedor,
                f.numero_factura                       AS nro_factura,
                f.descripcion,
                strftime('%d/%m/%Y', f.fecha_emision)    AS f_emision,
                strftime('%d/%m/%Y', f.fecha_vencimiento) AS f_vencimiento,
                f.estado,
                f.monto_total,
                f.monto_pagado,
                (f.monto_total - f.monto_pagado)       AS saldo,
                f.notas
            FROM facturas_proveedores f
            JOIN proveedores pr ON pr.id = f.proveedor_id
            ORDER BY pr.nombre, f.fecha_emision DESC
        """).fetchall()
        conn.close()

        encabezados = ["Proveedor", "N° Factura", "Descripción",
                       "F. Emisión", "F. Vencimiento", "Estado",
                       "Total", "Pagado", "Saldo", "Notas"]
        for col, h in enumerate(encabezados, 1):
            c = ws.cell(row=1, column=col, value=h)
            enc(c)
        ws.row_dimensions[1].height = 28

        nums = {7, 8, 9}
        for r, row in enumerate(rows, 2):
            vals = [
                row["proveedor"], row["nro_factura"] or "", row["descripcion"] or "",
                row["f_emision"], row["f_vencimiento"], row["estado"],
                row["monto_total"], row["monto_pagado"], row["saldo"],
                row["notas"] or "",
            ]
            saldo_pendiente = row["saldo"] > 0.01 and row["estado"] != "pagada"
            fill = PatternFill("solid", fgColor="FFD7D7") if saldo_pendiente \
                   else PatternFill("solid", fgColor=fila_par) if r % 2 == 0 \
                   else None
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = Alignment(horizontal="right" if col in nums else "left",
                                        vertical="center")
                if col in nums:
                    c.number_format = '#,##0.00'
                if fill:
                    c.fill = fill

        # Totales
        if rows:
            fila_tot = len(rows) + 2
            ws.cell(row=fila_tot, column=1, value="TOTALES")
            for i, col in enumerate([7, 8, 9]):
                c = ws.cell(row=fila_tot, column=col,
                            value=sum(r[col - 1] for r in
                                      [list(dict(row).values()) for row in rows]))
                c.number_format = '#,##0.00'
                tot(c)
            tot(ws.cell(row=fila_tot, column=1))

        ancho(ws)


# ─────────────────────────────────────────────────────────────
#  Worker para backup de la base de datos
# ─────────────────────────────────────────────────────────────

class BackupWorker(QThread):
    terminado = pyqtSignal(bool, str)   # (exito, ruta_o_error)

    def __init__(self, origen: str, destino: str):
        super().__init__()
        self.origen  = origen
        self.destino = destino

    def run(self):
        try:
            os.makedirs(os.path.dirname(self.destino), exist_ok=True)
            shutil.copy2(self.origen, self.destino)
            self.terminado.emit(True, self.destino)
        except Exception as e:
            self.terminado.emit(False, str(e))


# ─────────────────────────────────────────────────────────────
#  Panel principal de exportación
# ─────────────────────────────────────────────────────────────

class ExportarWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._worker = None
        self._backup_worker = None
        self._build_ui()
        # Verificar backup automático al iniciar (diferido 5s para no demorar el arranque)
        QTimer.singleShot(5000, self._verificar_backup_automatico)

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        # ── Scroll ────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        inner = QWidget()
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(32, 28, 32, 32)
        lay.setSpacing(20)
        scroll.setWidget(inner)
        outer.addWidget(scroll)

        # ── Título ────────────────────────────────────────────
        titulo = QLabel("📂  Exportar datos a Excel")
        titulo.setStyleSheet(
            "font-size:18px; font-weight:700; color:#C9A84C; padding-bottom:4px;")
        lay.addWidget(titulo)

        subtitulo = QLabel(
            "Descargá tus datos en formato Excel (.xlsx) cuando quieras.\n"
            "Tus registros de ventas, stock y cuentas siempre disponibles."
        )
        subtitulo.setStyleSheet("color:#AAAAAA; font-size:10pt;")
        lay.addWidget(subtitulo)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#333; margin: 4px 0;")
        lay.addWidget(sep)

        # ── Card: Ventas ──────────────────────────────────────
        lay.addWidget(self._card(
            titulo="🛒  Ventas",
            descripcion=(
                "Resumen diario (fecha, total, medio de pago, cantidad de ventas).\n"
                "Ideal para comparar con el registro histórico de Excel."
            ),
            tipo_resumen="ventas_resumen",
            tipo_detalle="ventas_detalle",
            con_periodo=True,
        ))

        # ── Card: Stock ───────────────────────────────────────
        lay.addWidget(self._card(
            titulo="📦  Stock actual",
            descripcion=(
                "Todos los productos activos con precios, stock actual,\n"
                "stock mínimo y alerta de reposición."
            ),
            tipo_resumen="stock",
            con_periodo=False,
        ))

        # ── Card: Cuentas proveedores ─────────────────────────
        lay.addWidget(self._card(
            titulo="🧾  Cuentas proveedores",
            descripcion=(
                "Todas las facturas de proveedores: vencimientos, estado,\n"
                "montos totales, pagados y saldo pendiente."
            ),
            tipo_resumen="proveedores",
            con_periodo=False,
        ))

        # ── Card: Exportación completa ────────────────────────
        lay.addWidget(self._card_completo())

        # ── Card: Backup automático ───────────────────────────
        lay.addWidget(self._card_backup())

        lay.addStretch()

    # ── Construcción de cards ──────────────────────────────────

    def _card(self, titulo, descripcion, tipo_resumen,
              tipo_detalle=None, con_periodo=False):
        frame = QFrame()
        frame.setObjectName("card_export")
        frame.setStyleSheet(
            "#card_export {"
            "  background:#1E1E1E; border:1px solid #333;"
            "  border-radius:8px; padding:16px;"
            "}"
        )
        lay = QVBoxLayout(frame)
        lay.setSpacing(10)

        lbl_tit = QLabel(titulo)
        lbl_tit.setStyleSheet("font-size:13px; font-weight:700; color:#DDDDDD;")
        lay.addWidget(lbl_tit)

        lbl_desc = QLabel(descripcion)
        lbl_desc.setStyleSheet("color:#888888; font-size:9pt;")
        lay.addWidget(lbl_desc)

        # Fila inferior: periodo (opcional) + botones
        fila = QHBoxLayout()
        fila.setSpacing(10)

        periodo_combo = None
        if con_periodo:
            periodo_combo = QComboBox()
            periodo_combo.addItems(["Histórico completo", "Este año",
                                    "Este mes", "Hoy"])
            periodo_combo.setFixedWidth(200)
            fila.addWidget(periodo_combo)

        fila.addStretch()

        # Botón resumen
        lbl_btn1 = "📊  Descargar resumen" if tipo_detalle else "📥  Descargar Excel"
        btn1 = QPushButton(lbl_btn1)
        btn1.setObjectName("btn_secundario")
        btn1.setFixedHeight(34)
        btn1.setMinimumWidth(180)
        btn1.clicked.connect(
            lambda _, t=tipo_resumen, pc=periodo_combo: self._exportar(t, pc)
        )
        fila.addWidget(btn1)

        if tipo_detalle:
            btn2 = QPushButton("📋  Descargar detalle")
            btn2.setObjectName("btn_secundario")
            btn2.setFixedHeight(34)
            btn2.setMinimumWidth(180)
            btn2.clicked.connect(
                lambda _, t=tipo_detalle, pc=periodo_combo: self._exportar(t, pc)
            )
            fila.addWidget(btn2)

        lay.addLayout(fila)
        return frame

    def _card_completo(self):
        frame = QFrame()
        frame.setObjectName("card_export_full")
        frame.setStyleSheet(
            "#card_export_full {"
            "  background:#2C1A1D; border:1px solid #722F37;"
            "  border-radius:8px; padding:16px;"
            "}"
        )
        lay = QVBoxLayout(frame)
        lay.setSpacing(10)

        lbl_tit = QLabel("📦  Exportación completa")
        lbl_tit.setStyleSheet("font-size:13px; font-weight:700; color:#C9A84C;")
        lay.addWidget(lbl_tit)

        lbl_desc = QLabel(
            "Un solo archivo con todas las hojas: ventas (resumen y detalle),\n"
            "stock actual y cuentas de proveedores."
        )
        lbl_desc.setStyleSheet("color:#888888; font-size:9pt;")
        lay.addWidget(lbl_desc)

        fila = QHBoxLayout()
        fila.addStretch()
        btn = QPushButton("💾  Descargar todo")
        btn.setObjectName("btn_exito")
        btn.setFixedHeight(36)
        btn.setMinimumWidth(200)
        btn.clicked.connect(lambda: self._exportar("completo", None))
        fila.addWidget(btn)
        lay.addLayout(fila)
        return frame

    # ── Lógica de exportación ──────────────────────────────────

    def _exportar(self, tipo: str, combo_periodo):
        PERIODOS = {
            "Histórico completo": "historico",
            "Este año":           "anio",
            "Este mes":           "mes",
            "Hoy":                "hoy",
        }
        periodo = "historico"
        if combo_periodo is not None:
            periodo = PERIODOS.get(combo_periodo.currentText(), "historico")

        NOMBRES = {
            "ventas_resumen": "ventas_resumen",
            "ventas_detalle": "ventas_detalle",
            "stock":          "stock_actual",
            "proveedores":    "cuentas_proveedores",
            "completo":       "vinoteca_completo",
        }
        hoy = datetime.today().strftime("%Y%m%d")
        nombre_ini = f"{NOMBRES[tipo]}_{hoy}.xlsx"

        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", nombre_ini,
            "Excel (*.xlsx)"
        )
        if not ruta:
            return

        self._worker = ExportWorker(tipo, ruta, periodo)
        self._worker.terminado.connect(self._on_terminado)
        self._worker.start()

    def _on_terminado(self, exito: bool, ruta_o_error: str):
        if exito:
            QMessageBox.information(
                self, "¡Listo!",
                f"✅  Archivo guardado correctamente:\n{ruta_o_error}"
            )
        else:
            QMessageBox.critical(
                self, "Error al exportar",
                f"No se pudo generar el archivo:\n{ruta_o_error}"
            )

    # ── Card: Backup automático ────────────────────────────────

    def _card_backup(self):
        frame = QFrame()
        frame.setObjectName("card_backup")
        frame.setStyleSheet(
            "#card_backup {"
            "  background:#1A2A1A; border:1px solid #2E7D32;"
            "  border-radius:8px; padding:16px;"
            "}"
        )
        lay = QVBoxLayout(frame)
        lay.setSpacing(12)

        lbl_tit = QLabel("🗄️  Backup automático")
        lbl_tit.setStyleSheet("font-size:13px; font-weight:700; color:#81C784;")
        lay.addWidget(lbl_tit)

        lbl_desc = QLabel(
            "Copia de seguridad de la base de datos (o exportación Excel selectiva).\n"
            "Se guarda en la carpeta que elijas (OneDrive, pendrive, etc.)."
        )
        lbl_desc.setStyleSheet("color:#888888; font-size:9pt;")
        lay.addWidget(lbl_desc)

        # ── Fila: carpeta destino ─────────────────────────────
        fila_carpeta = QHBoxLayout()
        fila_carpeta.setSpacing(8)
        lbl_c = QLabel("Carpeta:")
        lbl_c.setStyleSheet("color:#CCCCCC; min-width:70px;")
        fila_carpeta.addWidget(lbl_c)

        self._txt_carpeta = QLineEdit()
        self._txt_carpeta.setPlaceholderText("(ninguna seleccionada)")
        self._txt_carpeta.setReadOnly(True)
        self._txt_carpeta.setStyleSheet(
            "background:#111; border:1px solid #444; border-radius:4px;"
            "padding:4px 8px; color:#DDDDDD;"
        )
        carpeta_guardada = db.get_config("backup_carpeta", "")
        if carpeta_guardada:
            self._txt_carpeta.setText(carpeta_guardada)
        fila_carpeta.addWidget(self._txt_carpeta, 1)

        btn_elegir = QPushButton("📁  Elegir…")
        btn_elegir.setObjectName("btn_secundario")
        btn_elegir.setFixedHeight(30)
        btn_elegir.setFixedWidth(100)
        btn_elegir.clicked.connect(self._elegir_carpeta_backup)
        fila_carpeta.addWidget(btn_elegir)
        lay.addLayout(fila_carpeta)

        # ── Fila: modo de archivo (sobreescribir / con fecha) ──
        fila_modo = QHBoxLayout()
        fila_modo.setSpacing(8)
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        lbl_modo = QLabel("Modo:")
        lbl_modo.setStyleSheet("color:#CCCCCC; min-width:70px;")
        fila_modo.addWidget(lbl_modo)

        self._grp_modo = QButtonGroup(frame)
        self._rb_sobreescribir = QRadioButton("Sobreescribir (un solo archivo)")
        self._rb_sobreescribir.setStyleSheet("color:#CCCCCC; font-size:9pt;")
        self._rb_con_fecha = QRadioButton("Agregar con fecha (historial)")
        self._rb_con_fecha.setStyleSheet("color:#CCCCCC; font-size:9pt;")
        self._grp_modo.addButton(self._rb_sobreescribir, 0)
        self._grp_modo.addButton(self._rb_con_fecha, 1)

        modo_guardado = db.get_config("backup_modo", "fecha")
        if modo_guardado == "sobreescribir":
            self._rb_sobreescribir.setChecked(True)
        else:
            self._rb_con_fecha.setChecked(True)

        def _guardar_modo():
            modo = "sobreescribir" if self._rb_sobreescribir.isChecked() else "fecha"
            db.set_config("backup_modo", modo, "string")

        self._rb_sobreescribir.toggled.connect(_guardar_modo)

        fila_modo.addWidget(self._rb_sobreescribir)
        fila_modo.addSpacing(16)
        fila_modo.addWidget(self._rb_con_fecha)
        fila_modo.addStretch()
        lay.addLayout(fila_modo)

        # ── Fila: frecuencia + último backup ──────────────────
        fila_freq = QHBoxLayout()
        fila_freq.setSpacing(8)
        lbl_f = QLabel("Cada:")
        lbl_f.setStyleSheet("color:#CCCCCC; min-width:70px;")
        fila_freq.addWidget(lbl_f)

        self._spin_dias = QSpinBox()
        self._spin_dias.setRange(1, 365)
        self._spin_dias.setSuffix(" días")
        self._spin_dias.setFixedWidth(100)
        self._spin_dias.setStyleSheet(
            "QSpinBox{background:#111;border:1px solid #444;border-radius:4px;"
            "padding:2px 22px 2px 6px;color:#DDDDDD;}"
            + _SPIN_SUBS
        )
        freq_guardada = db.get_config("backup_frecuencia_dias", 7)
        try:
            self._spin_dias.setValue(int(freq_guardada))
        except Exception:
            self._spin_dias.setValue(7)
        self._spin_dias.valueChanged.connect(
            lambda v: db.set_config("backup_frecuencia_dias", v, "int")
        )
        fila_freq.addWidget(self._spin_dias)

        fila_freq.addSpacing(20)

        ultima = db.get_config("backup_ultima_fecha", "")
        txt_ultima = f"Último backup: {ultima}" if ultima else "Nunca se hizo backup"
        self._lbl_ultima = QLabel(txt_ultima)
        self._lbl_ultima.setStyleSheet("color:#666666; font-size:9pt;")
        fila_freq.addWidget(self._lbl_ultima)
        fila_freq.addStretch()
        lay.addLayout(fila_freq)

        # ── Qué incluir en el backup ──────────────────────────
        sep_scope = QFrame()
        sep_scope.setFrameShape(QFrame.Shape.HLine)
        sep_scope.setStyleSheet("color:#2E7D32; margin:2px 0;")
        lay.addWidget(sep_scope)

        lbl_scope = QLabel("¿Qué incluir?")
        lbl_scope.setStyleSheet("color:#CCCCCC; font-size:9pt; font-weight:600;")
        lay.addWidget(lbl_scope)

        from PyQt6.QtWidgets import QCheckBox as _QChk
        fila_scope = QHBoxLayout()
        fila_scope.setSpacing(16)

        chk_style = "color:#CCCCCC; font-size:9pt;"
        self._chk_bk_db       = _QChk("🗄 Base de datos completa")
        self._chk_bk_ventas   = _QChk("🛒 Excel ventas")
        self._chk_bk_stock    = _QChk("📦 Excel stock")
        self._chk_bk_cuentas  = _QChk("🧾 Excel cuentas")

        for chk in (self._chk_bk_db, self._chk_bk_ventas,
                    self._chk_bk_stock, self._chk_bk_cuentas):
            chk.setStyleSheet(chk_style)

        # Restaurar preferencias guardadas
        self._chk_bk_db.setChecked(db.get_config("backup_scope_db", "1") != "0")
        self._chk_bk_ventas.setChecked(db.get_config("backup_scope_ventas", "0") == "1")
        self._chk_bk_stock.setChecked(db.get_config("backup_scope_stock", "0") == "1")
        self._chk_bk_cuentas.setChecked(db.get_config("backup_scope_cuentas", "0") == "1")

        def _save_scope():
            db.set_config("backup_scope_db",      "1" if self._chk_bk_db.isChecked()      else "0", "string")
            db.set_config("backup_scope_ventas",   "1" if self._chk_bk_ventas.isChecked()  else "0", "string")
            db.set_config("backup_scope_stock",    "1" if self._chk_bk_stock.isChecked()   else "0", "string")
            db.set_config("backup_scope_cuentas",  "1" if self._chk_bk_cuentas.isChecked() else "0", "string")

        for chk in (self._chk_bk_db, self._chk_bk_ventas,
                    self._chk_bk_stock, self._chk_bk_cuentas):
            chk.stateChanged.connect(lambda _: _save_scope())

        fila_scope.addWidget(self._chk_bk_db)
        fila_scope.addWidget(self._chk_bk_ventas)
        fila_scope.addWidget(self._chk_bk_stock)
        fila_scope.addWidget(self._chk_bk_cuentas)
        fila_scope.addStretch()
        lay.addLayout(fila_scope)

        # ── Botón manual ──────────────────────────────────────
        fila_btn = QHBoxLayout()
        fila_btn.addStretch()
        self._btn_backup_ahora = QPushButton("💾  Hacer backup ahora")
        self._btn_backup_ahora.setObjectName("btn_exito")
        self._btn_backup_ahora.setFixedHeight(34)
        self._btn_backup_ahora.setMinimumWidth(200)
        self._btn_backup_ahora.clicked.connect(self._hacer_backup)
        fila_btn.addWidget(self._btn_backup_ahora)
        lay.addLayout(fila_btn)

        return frame

    def _elegir_carpeta_backup(self):
        carpeta = QFileDialog.getExistingDirectory(
            self, "Elegir carpeta de backup",
            self._txt_carpeta.text() or os.path.expanduser("~"),
        )
        if carpeta:
            self._txt_carpeta.setText(carpeta)
            db.set_config("backup_carpeta", carpeta, "string")

    def _hacer_backup(self, silencioso=False):
        carpeta = self._txt_carpeta.text().strip()
        if not carpeta:
            if not silencioso:
                QMessageBox.warning(
                    self, "Sin carpeta",
                    "Primero elegí una carpeta de destino para el backup."
                )
            return

        # Verificar que al menos un scope esté seleccionado
        include_db      = hasattr(self, "_chk_bk_db")      and self._chk_bk_db.isChecked()
        include_ventas  = hasattr(self, "_chk_bk_ventas")  and self._chk_bk_ventas.isChecked()
        include_stock   = hasattr(self, "_chk_bk_stock")   and self._chk_bk_stock.isChecked()
        include_cuentas = hasattr(self, "_chk_bk_cuentas") and self._chk_bk_cuentas.isChecked()

        if not any([include_db, include_ventas, include_stock, include_cuentas]):
            if not silencioso:
                QMessageBox.warning(self, "Nada seleccionado",
                    "Marcá al menos una opción de qué incluir en el backup.")
            return

        sobreescribir = (hasattr(self, "_rb_sobreescribir")
                         and self._rb_sobreescribir.isChecked())

        fecha_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        self._btn_backup_ahora.setEnabled(False)
        self._btn_backup_ahora.setText("Guardando…")

        self._pending_backups = []          # lista de (ok, ruta)
        self._total_backups   = 0

        # ── 1. Copia de la base de datos ──────────────────────
        if include_db:
            from config import BASE_DIR
            origen = os.path.join(BASE_DIR, "db", "vinoteca.db")
            if os.path.exists(origen):
                if sobreescribir:
                    nombre_db = "vinoteca_backup.db"
                else:
                    nombre_db = f"vinoteca_backup_{fecha_str}.db"
                destino_db = os.path.join(carpeta, nombre_db)
                self._total_backups += 1
                _w = BackupWorker(origen, destino_db)
                _w.terminado.connect(self._on_un_backup_done)
                _w.start()
                self._backup_worker = _w   # evitar GC

        # ── 2. Excel por tipo ─────────────────────────────────
        TIPOS_EXCEL = []
        if include_ventas:
            TIPOS_EXCEL.append(("completo_ventas", "ventas_completo"))
        if include_stock:
            TIPOS_EXCEL.append(("stock", "stock_actual"))
        if include_cuentas:
            TIPOS_EXCEL.append(("proveedores", "cuentas_proveedores"))

        self._excel_workers = []
        for tipo_worker, nombre_base in TIPOS_EXCEL:
            if sobreescribir:
                nombre_xlsx = f"{nombre_base}.xlsx"
            else:
                nombre_xlsx = f"{nombre_base}_{fecha_str}.xlsx"
            ruta_xlsx = os.path.join(carpeta, nombre_xlsx)

            # Para ventas usamos reporte completo (resumen + detalle)
            tipo_real = tipo_worker
            if tipo_worker == "completo_ventas":
                tipo_real = "completo"

            self._total_backups += 1
            _ew = ExportWorker(tipo_real, ruta_xlsx, "historico")
            _ew.terminado.connect(self._on_un_backup_done)
            _ew.start()
            self._excel_workers.append(_ew)

        if self._total_backups == 0:
            self._btn_backup_ahora.setEnabled(True)
            self._btn_backup_ahora.setText("💾  Hacer backup ahora")

    def _on_un_backup_done(self, exito: bool, ruta_o_error: str):
        self._pending_backups.append((exito, ruta_o_error))
        if len(self._pending_backups) < self._total_backups:
            return   # esperar el resto

        # Todos terminaron
        self._btn_backup_ahora.setEnabled(True)
        self._btn_backup_ahora.setText("💾  Hacer backup ahora")

        errores  = [msg for ok, msg in self._pending_backups if not ok]
        correctos = [ruta for ok, ruta in self._pending_backups if ok]

        if correctos:
            ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
            db.set_config("backup_ultima_fecha", ahora, "string")
            self._lbl_ultima.setText(f"Último backup: {ahora}")

        if errores:
            QMessageBox.critical(
                self, "Error en backup",
                f"Algunos archivos no se pudieron guardar:\n" + "\n".join(errores)
            )
        elif correctos:
            QMessageBox.information(
                self, "Backup exitoso",
                f"✅  {len(correctos)} archivo(s) guardado(s) correctamente en:\n"
                + os.path.dirname(correctos[0])
            )

    def _on_backup_terminado(self, exito: bool, ruta_o_error: str, silencioso: bool):
        """Compatibilidad con backups silenciosos del auto-backup."""
        self._btn_backup_ahora.setEnabled(True)
        self._btn_backup_ahora.setText("💾  Hacer backup ahora")

        if exito:
            ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
            db.set_config("backup_ultima_fecha", ahora, "string")
            self._lbl_ultima.setText(f"Último backup: {ahora}")
            if not silencioso:
                QMessageBox.information(
                    self, "Backup exitoso",
                    f"✅  Backup guardado correctamente:\n{ruta_o_error}"
                )
        else:
            if not silencioso:
                QMessageBox.critical(
                    self, "Error en backup",
                    f"No se pudo guardar el backup:\n{ruta_o_error}"
                )

    def _verificar_backup_automatico(self):
        """Comprueba si corresponde hacer un backup automático según la frecuencia configurada."""
        carpeta = db.get_config("backup_carpeta", "")
        if not carpeta or not os.path.isdir(carpeta):
            return

        frecuencia = db.get_config("backup_frecuencia_dias", 7)
        ultima = db.get_config("backup_ultima_fecha", "")

        if ultima:
            try:
                # El formato guardado es "dd/mm/YYYY HH:MM"
                ultima_dt = datetime.strptime(ultima, "%d/%m/%Y %H:%M")
                dias_pasados = (datetime.now() - ultima_dt).days
                if dias_pasados < int(frecuencia):
                    return   # No toca todavía
            except Exception:
                pass   # Si hay error de parseo, hacemos backup igual

        # Corresponde hacer backup silencioso
        self._hacer_backup(silencioso=True)

