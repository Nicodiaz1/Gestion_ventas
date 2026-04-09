# ─────────────────────────────────────────────────────────────
#  ui/stock.py  –  Gestión de Stock y Productos
# ─────────────────────────────────────────────────────────────

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QDialog, QFormLayout, QComboBox, QSpinBox,
    QDoubleSpinBox, QAbstractItemView, QTabWidget, QTextEdit,
    QFrame, QCheckBox, QSizePolicy, QDateEdit, QStyledItemDelegate, QMenu
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QDate
from PyQt6.QtGui import QColor, QFont
import sys, os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db import database as db


class FilaColorDelegate(QStyledItemDelegate):
    """Fuerza el background personalizado incluso con QSS activo en macOS."""
    def paint(self, painter, option, index):
        from PyQt6.QtGui import QBrush, QPalette
        from PyQt6.QtWidgets import QStyle, QStyleOptionViewItem

        bg = index.data(Qt.ItemDataRole.BackgroundRole)
        is_selected = bool(option.state & QStyle.StateFlag.State_Selected)

        if bg is not None and not is_selected:
            opt = QStyleOptionViewItem(option)
            self.initStyleOption(opt, index)

            # Obtener el color del brush
            color = bg if isinstance(bg, QColor) else bg.color()

            # Parchar la paleta con nuestro color → QMacStyle y QCommonStyle lo usan
            palette = QPalette(opt.palette)
            palette.setColor(QPalette.ColorRole.Base, color)
            palette.setColor(QPalette.ColorRole.AlternateBase, color)
            opt.palette = palette
            opt.backgroundBrush = QBrush(color)

            # Pre-rellenar el rect nosotros también (doble seguro)
            painter.save()
            painter.fillRect(option.rect, color)
            painter.restore()

            super().paint(painter, opt, index)
        else:
            super().paint(painter, option, index)


# ─────────────────────────────────────────────────────────────
#  Diálogo: Crear / Editar Producto
# ─────────────────────────────────────────────────────────────

class DialogoProducto(QDialog):
    def __init__(self, parent=None, producto: dict = None):
        super().__init__(parent)
        self.producto = producto
        self.setWindowTitle("Nuevo Producto" if not producto else "Editar Producto")
        self.setMinimumSize(600, 560)
        self.setModal(True)
        self._build_ui()
        if producto:
            self._cargar_datos()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(12)
        lay.setContentsMargins(24, 20, 24, 20)

        titulo = QLabel("📦  " + self.windowTitle())
        titulo.setObjectName("titulo_seccion")
        lay.addWidget(titulo)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)

        # Nombre
        self.txt_nombre = QLineEdit()
        self.txt_nombre.setPlaceholderText("Ej: Malbec Reserva 750ml")
        form.addRow("Nombre *:", self.txt_nombre)

        # Código de barras
        codigo_row = QHBoxLayout()
        self.txt_codigo = QLineEdit()
        self.txt_codigo.setPlaceholderText("Escaneá o escribí el código")
        self.chk_sin_codigo = QCheckBox("Sin código de barras")
        self.chk_sin_codigo.stateChanged.connect(self._toggle_codigo)
        codigo_row.addWidget(self.txt_codigo, 1)
        codigo_row.addWidget(self.chk_sin_codigo)
        form.addRow("Código de barras:", codigo_row)

        # Categoría
        self.cmb_categoria = QComboBox()
        self._cargar_categorias()
        form.addRow("Categoría:", self.cmb_categoria)

        # Precio de venta
        self.spin_precio = QDoubleSpinBox()
        self.spin_precio.setRange(0, 9999999)
        self.spin_precio.setDecimals(2)
        self.spin_precio.setSingleStep(100)
        self.spin_precio.setPrefix("$ ")
        form.addRow("Precio de venta *:", self.spin_precio)

        # Precio de costo
        self.spin_costo = QDoubleSpinBox()
        self.spin_costo.setRange(0, 9999999)
        self.spin_costo.setDecimals(2)
        self.spin_costo.setSingleStep(100)
        self.spin_costo.setPrefix("$ ")
        form.addRow("Precio de costo:", self.spin_costo)

        # Stock inicial
        self.spin_stock = QSpinBox()
        self.spin_stock.setRange(0, 99999)
        self.spin_stock.setSingleStep(1)
        form.addRow("Stock inicial:", self.spin_stock)

        # Stock mínimo (alerta)
        self.spin_stock_min = QSpinBox()
        self.spin_stock_min.setRange(0, 9999)
        self.spin_stock_min.setValue(3)
        form.addRow("Alerta de stock mínimo:", self.spin_stock_min)

        # Unidad
        self.cmb_unidad = QComboBox()
        for u in ["botella", "caja", "lata", "sixpack", "unidad", "pack", "otro"]:
            self.cmb_unidad.addItem(u)
        form.addRow("Unidad de venta:", self.cmb_unidad)

        # Unidades por caja
        caja_row = QHBoxLayout()
        self.spin_upc = QSpinBox()
        self.spin_upc.setRange(1, 999)
        self.spin_upc.setValue(1)
        self.spin_upc.setSingleStep(1)
        self.lbl_upc_hint = QLabel("unidades por caja  →  En la carga de stock podés ingresar por caja")
        self.lbl_upc_hint.setStyleSheet("color:#888; font-size:9pt;")
        caja_row.addWidget(self.spin_upc)
        caja_row.addWidget(self.lbl_upc_hint)
        caja_row.addStretch()
        form.addRow("Unidades por caja:", caja_row)
        self.spin_upc.valueChanged.connect(self._actualizar_hint_caja)

        # Descripción
        self.txt_descripcion = QTextEdit()
        self.txt_descripcion.setMaximumHeight(80)
        self.txt_descripcion.setPlaceholderText("Notas opcionales sobre el producto…")
        form.addRow("Descripción:", self.txt_descripcion)

        # Separador visual
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#333;")
        form.addRow(sep)

        # Año de cosecha
        cosecha_row = QHBoxLayout()
        self.spin_cosecha = QSpinBox()
        self.spin_cosecha.setRange(0, 2100)
        self.spin_cosecha.setSpecialValueText("Sin especificar")
        self.spin_cosecha.setValue(0)
        self.spin_cosecha.setToolTip("Año de cosecha / añada del vino (0 = sin especificar)")
        lbl_cosecha_hint = QLabel("(0 = sin especificar)")
        lbl_cosecha_hint.setStyleSheet("color:#666; font-size:9pt;")
        cosecha_row.addWidget(self.spin_cosecha)
        cosecha_row.addWidget(lbl_cosecha_hint)
        cosecha_row.addStretch()
        form.addRow("🍇  Año de cosecha:", cosecha_row)

        # Fecha de vencimiento: checkbox en su propia fila, datepicker en la siguiente
        self.chk_tiene_venc = QCheckBox("Tiene fecha de vencimiento")
        form.addRow("📅  Vencimiento:", self.chk_tiene_venc)

        date_row = QHBoxLayout()
        self.date_venc = QDateEdit()
        self.date_venc.setCalendarPopup(True)
        self.date_venc.setDisplayFormat("dd/MM/yyyy")
        self.date_venc.setDate(QDate.currentDate().addMonths(6))
        self.date_venc.setVisible(False)
        date_row.addWidget(self.date_venc)
        date_row.addStretch()
        form.addRow("", date_row)
        self.chk_tiene_venc.stateChanged.connect(
            lambda s: self.date_venc.setVisible(s == Qt.CheckState.Checked.value))

        lay.addLayout(form)

        # Botones
        btn_row = QHBoxLayout()
        btn_ok = QPushButton("💾  Guardar")
        btn_ok.clicked.connect(self._guardar)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

    def _actualizar_hint_caja(self, valor: int):
        if valor <= 1:
            self.lbl_upc_hint.setText("unidad por unidad  (sin agrupamiento en cajas)")
        else:
            self.lbl_upc_hint.setText(f"{valor} unidades por caja  →  En la carga podés ingresar por caja")

    def _cargar_categorias(self):
        self.cmb_categoria.clear()
        self.cmb_categoria.addItem("— Sin categoría —", None)
        for c in db.obtener_categorias():
            self.cmb_categoria.addItem(c["nombre"], c["id"])

    def _toggle_codigo(self, state):
        self.txt_codigo.setEnabled(state != Qt.CheckState.Checked.value)
        if state == Qt.CheckState.Checked.value:
            self.txt_codigo.clear()

    def _cargar_datos(self):
        p = self.producto
        self.txt_nombre.setText(p.get("nombre", ""))
        codigo = p.get("codigo_barras") or ""
        self.txt_codigo.setText(codigo)
        if p.get("sin_codigo"):
            self.chk_sin_codigo.setChecked(True)
        self.spin_precio.setValue(p.get("precio_venta", 0))
        self.spin_costo.setValue(p.get("precio_costo", 0))
        self.spin_stock.setValue(p.get("stock_actual", 0))
        self.spin_stock_min.setValue(p.get("stock_minimo", 3))
        self.spin_upc.setValue(p.get("unidades_por_caja") or 1)
        self.txt_descripcion.setPlainText(p.get("descripcion") or "")

        # Unidad
        idx = self.cmb_unidad.findText(p.get("unidad", "botella"))
        if idx >= 0:
            self.cmb_unidad.setCurrentIndex(idx)

        # Categoría
        cat_id = p.get("categoria_id")
        for i in range(self.cmb_categoria.count()):
            if self.cmb_categoria.itemData(i) == cat_id:
                self.cmb_categoria.setCurrentIndex(i)
                break

        # Cargar cosecha y vencimiento del lote más reciente
        lotes = db.obtener_lotes_producto(p["id"])
        if lotes:
            lote0 = lotes[-1]   # más reciente (orden FIFO → último tiene cosecha más nueva)
            if lote0["cosecha"]:
                self.spin_cosecha.setValue(lote0["cosecha"])
            SENTINEL = "2000-01-01"
            venc = lote0["fecha_vencimiento"]
            if venc and str(venc)[:10] != SENTINEL:
                self.chk_tiene_venc.setChecked(True)
                self.date_venc.setDate(QDate.fromString(str(venc)[:10], "yyyy-MM-dd"))

    def _guardar(self):
        nombre = self.txt_nombre.text().strip()
        if not nombre:
            QMessageBox.warning(self, "Campo requerido", "El nombre del producto es obligatorio.")
            return

        sin_codigo = self.chk_sin_codigo.isChecked()
        codigo = self.txt_codigo.text().strip() or None
        if not sin_codigo and not codigo:
            resp = QMessageBox.question(
                self, "Sin código",
                "No ingresaste código de barras.\n¿Deseas marcarlo como producto sin código?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if resp == QMessageBox.StandardButton.Yes:
                sin_codigo = True
            else:
                return

        datos = {
            "nombre": nombre,
            "codigo_barras": None if sin_codigo else codigo,
            "sin_codigo": 1 if sin_codigo else 0,
            "precio_venta": self.spin_precio.value(),
            "precio_costo": self.spin_costo.value(),
            "stock_actual": self.spin_stock.value(),
            "stock_minimo": self.spin_stock_min.value(),
            "unidad": self.cmb_unidad.currentText(),
            "unidades_por_caja": self.spin_upc.value(),
            "descripcion": self.txt_descripcion.toPlainText().strip(),
            "categoria_id": self.cmb_categoria.currentData(),
        }

        try:
            cosecha = self.spin_cosecha.value() or None
            fecha_venc = (self.date_venc.date().toString("yyyy-MM-dd")
                          if self.chk_tiene_venc.isChecked() else None)
            if self.producto:
                db.actualizar_producto(self.producto["id"], datos)
                # Actualizar cosecha/vencimiento en el lote más reciente
                lotes = db.obtener_lotes_producto(self.producto["id"])
                if lotes:
                    db.actualizar_lote(lotes[-1]["id"],
                                       cosecha=cosecha,
                                       fecha_vencimiento=fecha_venc)
                elif (cosecha or fecha_venc) and datos["stock_actual"] > 0:
                    # No había lotes — crear uno con el stock actual
                    db.crear_lote(self.producto["id"], datos["stock_actual"],
                                  fecha_vencimiento=fecha_venc,
                                  cosecha=cosecha,
                                  motivo="Lote creado al editar producto")
            else:
                pid = db.crear_producto(datos)
                # Si se especificó cosecha o vencimiento al crear, registrar lote inicial
                if (cosecha or fecha_venc) and datos["stock_actual"] > 0:
                    db.crear_lote(pid, datos["stock_actual"],
                                  fecha_vencimiento=fecha_venc,
                                  cosecha=cosecha,
                                  motivo="Ingreso inicial")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))

    def get_datos(self) -> dict:
        return {}


# ─────────────────────────────────────────────────────────────
#  Diálogo: Carga rápida de stock por escaneo
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
#  Diálogo: Ajuste manual de stock (rotura, robo, merma…)
# ─────────────────────────────────────────────────────────────

class DialogoAjusteStock(QDialog):
    def __init__(self, parent, producto: dict):
        super().__init__(parent)
        self.producto = producto
        self.setWindowTitle(f"⚖️  Ajuste de stock — {producto['nombre']}")
        self.setMinimumWidth(400)
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(12)
        lay.setContentsMargins(24, 20, 24, 20)

        lbl_nombre = QLabel(f"<b>{self.producto['nombre']}</b>")
        lbl_nombre.setStyleSheet("font-size:13pt;")
        lay.addWidget(lbl_nombre)

        self.lbl_stock = QLabel(f"Stock actual: <b>{self.producto['stock_actual']}</b>")
        self.lbl_stock.setStyleSheet("color:#AAAAAA;")
        lay.addWidget(self.lbl_stock)

        linea = QFrame()
        linea.setFrameShape(QFrame.Shape.HLine)
        linea.setStyleSheet("color:#333;")
        lay.addWidget(linea)

        form = QFormLayout()
        form.setSpacing(10)

        self.cmb_motivo = QComboBox()
        self.cmb_motivo.addItems([
            "Rotura", "Robo / Hurto", "Vencimiento / Merma",
            "Corrección de inventario", "Donación", "Otro"
        ])
        form.addRow("Motivo del ajuste:", self.cmb_motivo)

        self.spin_nuevo = QSpinBox()
        self.spin_nuevo.setRange(0, 99999)
        self.spin_nuevo.setValue(self.producto["stock_actual"])
        self.spin_nuevo.valueChanged.connect(self._actualizar_diferencia)
        form.addRow("Nuevo stock a fijar:", self.spin_nuevo)

        self.lbl_diferencia = QLabel("")
        form.addRow("Diferencia:", self.lbl_diferencia)

        self.txt_notas = QLineEdit()
        self.txt_notas.setPlaceholderText("Detalle adicional (opcional)")
        form.addRow("Notas:", self.txt_notas)

        lay.addLayout(form)

        btn_row = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = QPushButton("✅  Confirmar ajuste")
        btn_ok.setStyleSheet(
            "QPushButton{background:#E65100;color:white;font-weight:700;"
            "border-radius:6px;padding:8px 20px;}"
            "QPushButton:hover{background:#F57C00;}"
        )
        btn_ok.clicked.connect(self.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        self._actualizar_diferencia()

    def _actualizar_diferencia(self):
        diff = self.spin_nuevo.value() - self.producto["stock_actual"]
        if diff < 0:
            self.lbl_diferencia.setText(
                f"<b style='color:#EF5350'>{diff} unidades (reducción)</b>")
        elif diff > 0:
            self.lbl_diferencia.setText(
                f"<b style='color:#66BB6A'>+{diff} unidades (aumento)</b>")
        else:
            self.lbl_diferencia.setText("Sin cambio")

    def get_values(self) -> tuple:
        """Retorna (nuevo_stock, motivo_con_notas)."""
        motivo = self.cmb_motivo.currentText()
        notas  = self.txt_notas.text().strip()
        return self.spin_nuevo.value(), f"{motivo}{': ' + notas if notas else ''}"


# ─────────────────────────────────────────────────────────────
#  Diálogo: Ver lotes / cosechas de un producto
# ─────────────────────────────────────────────────────────────

class DialogoLotesProducto(QDialog):
    def __init__(self, parent, producto: dict):
        super().__init__(parent)
        self.producto = producto
        self.setWindowTitle(f"📋  Lotes / Cosechas — {producto['nombre']}")
        self.setMinimumSize(620, 380)
        self.setModal(True)
        self._build_ui()
        self._cargar()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 16, 20, 16)

        lbl = QLabel(f"<b>{self.producto['nombre']}</b>"
                     f"  —  Stock total: {self.producto['stock_actual']}")
        lbl.setStyleSheet("font-size:12pt; color:#C9A84C;")
        lay.addWidget(lbl)

        self.tabla = QTableWidget(0, 5)
        self.tabla.setHorizontalHeaderLabels(
            ["Cosecha", "Vencimiento", "Cantidad", "Fecha ingreso", "Motivo"])
        hh = self.tabla.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay.addWidget(self.tabla)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("Cerrar")
        btn_ok.clicked.connect(self.accept)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

    def _cargar(self):
        from datetime import date
        hoy   = date.today()
        lotes = db.obtener_lotes_producto(self.producto["id"])

        stock_total    = self.producto["stock_actual"] or 0
        stock_en_lotes = sum(l["cantidad"] for l in lotes)
        stock_sin_lote = max(0, stock_total - stock_en_lotes)

        # Filas: lotes + (fila sin lote si corresponde) + fila TOTAL
        n_filas = len(lotes) + (1 if stock_sin_lote > 0 else 0) + 1
        self.tabla.setRowCount(n_filas)

        for i, l in enumerate(lotes):
            cosecha = str(l["cosecha"]) if l["cosecha"] else "—"
            self.tabla.setItem(i, 0, QTableWidgetItem(cosecha))

            venc_str = l["fecha_vencimiento"]
            SENTINEL = "2000-01-01"
            if venc_str and venc_str[:10] != SENTINEL:
                try:
                    dias = (date.fromisoformat(venc_str[:10]) - hoy).days
                except Exception:
                    dias = None
                if dias is not None:
                    venc_item = QTableWidgetItem(f"{venc_str[:10]}  ({'+' if dias >= 0 else ''}{dias}d)")
                    if dias < 0:
                        venc_item.setForeground(QColor("#9E9E9E"))
                    elif dias <= 7:
                        venc_item.setForeground(QColor("#EF5350"))
                    elif dias <= 30:
                        venc_item.setForeground(QColor("#FF9800"))
                    else:
                        venc_item.setForeground(QColor("#66BB6A"))
                else:
                    venc_item = QTableWidgetItem(venc_str[:10])
            else:
                venc_item = QTableWidgetItem("—")
                venc_item.setForeground(QColor("#666"))
            self.tabla.setItem(i, 1, venc_item)

            cant_item = QTableWidgetItem(str(l["cantidad"]))
            cant_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla.setItem(i, 2, cant_item)
            self.tabla.setItem(i, 3, QTableWidgetItem(str(l["fecha_ingreso"])[:10]))
            self.tabla.setItem(i, 4, QTableWidgetItem(l["motivo"] or ""))
            self.tabla.setRowHeight(i, 36)

        # Fila "Sin lote" si hay stock no asignado a ningún lote
        idx = len(lotes)
        if stock_sin_lote > 0:
            def _gray(txt, col):
                it = QTableWidgetItem(txt)
                it.setForeground(QColor("#AAAAAA"))
                if col == 2:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                return it
            self.tabla.setItem(idx, 0, _gray("Sin lote / cosecha", 0))
            self.tabla.setItem(idx, 1, _gray("—", 1))
            cant_sl = QTableWidgetItem(str(stock_sin_lote))
            cant_sl.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            cant_sl.setForeground(QColor("#AAAAAA"))
            self.tabla.setItem(idx, 2, cant_sl)
            self.tabla.setItem(idx, 3, _gray("—", 3))
            self.tabla.setItem(idx, 4, _gray("Stock sin trazabilidad", 4))
            self.tabla.setRowHeight(idx, 36)
            idx += 1

        # Fila TOTAL
        total_item = QTableWidgetItem("TOTAL")
        total_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        total_item.setForeground(QColor("#C9A84C"))
        self.tabla.setItem(idx, 0, total_item)
        self.tabla.setItem(idx, 1, QTableWidgetItem(""))
        cant_total = QTableWidgetItem(str(stock_total))
        cant_total.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        cant_total.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        cant_total.setForeground(QColor("#C9A84C"))
        self.tabla.setItem(idx, 2, cant_total)
        self.tabla.setItem(idx, 3, QTableWidgetItem(""))
        self.tabla.setItem(idx, 4, QTableWidgetItem(""))
        self.tabla.setRowHeight(idx, 36)

        if not lotes and stock_sin_lote == 0:
            self.tabla.setRowCount(1)
            item = QTableWidgetItem("No hay stock registrado para este producto")
            item.setForeground(QColor("#666"))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla.setItem(0, 0, item)
            self.tabla.setSpan(0, 0, 1, 5)


# ─────────────────────────────────────────────────────────────
#  Diálogo: Carga de stock
# ─────────────────────────────────────────────────────────────

class DialogoCargaStock(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📥  Carga de Stock")
        self.setMinimumSize(820, 520)
        self.setModal(True)
        self.items: dict = {}    # producto_id → {"producto", "cantidad", "modo"}
        self._build_ui()
        QTimer.singleShot(100, self.scan_input.setFocus)

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 20, 20, 20)

        titulo = QLabel("📥  Ingreso de Mercadería")
        titulo.setObjectName("titulo_seccion")
        lay.addWidget(titulo)

        hint = QLabel("💡  Escaneá cada producto. Si tiene cajas configuradas podés elegir ingresar por caja.")
        hint.setStyleSheet("color:#AAAAAA; font-size:10pt;")
        lay.addWidget(hint)

        scan_row = QHBoxLayout()
        self.scan_input = QLineEdit()
        self.scan_input.setObjectName("scan_input")
        self.scan_input.setPlaceholderText("📷  Escaneá el código de barras…")
        self.scan_input.returnPressed.connect(self._procesar_escaneo)
        scan_row.addWidget(self.scan_input, 1)

        btn_buscar = QPushButton("🔍  Buscar")
        btn_buscar.clicked.connect(self._abrir_buscador)
        scan_row.addWidget(btn_buscar)
        lay.addLayout(scan_row)

        # Tabla de carga — 7 columnas
        self.tabla = QTableWidget(0, 8)
        self.tabla.setHorizontalHeaderLabels([
            "Código", "Producto", "Stock\nactual", "Modo de ingreso",
            "Cantidad", "Total\nunidades", "Cosecha\n(año)", "Fecha\nvencim."
        ])
        self.tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tabla.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla.setColumnWidth(0, 100)
        self.tabla.setColumnWidth(2, 70)
        self.tabla.setColumnWidth(4, 90)
        self.tabla.setColumnWidth(5, 110)
        self.tabla.setColumnWidth(6, 90)
        self.tabla.setColumnWidth(7, 120)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.verticalHeader().setVisible(False)
        lay.addWidget(self.tabla, 1)

        # Fila de resumen + motivo
        resumen_row = QHBoxLayout()
        self.lbl_resumen = QLabel("")
        self.lbl_resumen.setStyleSheet("color:#C9A84C; font-weight:700; font-size:10pt;")
        resumen_row.addWidget(self.lbl_resumen)
        resumen_row.addStretch()
        resumen_row.addWidget(QLabel("Motivo:"))
        self.txt_motivo = QLineEdit("Ingreso de mercadería")
        self.txt_motivo.setMaximumWidth(280)
        resumen_row.addWidget(self.txt_motivo)
        lay.addLayout(resumen_row)

        # Botones
        btn_row = QHBoxLayout()
        btn_ok = QPushButton("✅  Confirmar ingreso")
        btn_ok.setStyleSheet(
            "QPushButton{background:#2E7D32;font-size:12pt;font-weight:700;"
            "color:white;border-radius:8px;padding:10px 24px;}"
            "QPushButton:hover{background:#388E3C;}"
        )
        btn_ok.clicked.connect(self._confirmar)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

    def _procesar_escaneo(self):
        codigo = self.scan_input.text().strip()
        if not codigo:
            return
        self.scan_input.clear()
        producto = db.buscar_por_codigo(codigo)
        if producto:
            self._agregar_producto(dict(producto))
        else:
            QMessageBox.warning(self, "No encontrado",
                                f"Código '{codigo}' no existe.\n"
                                "Primero dá de alta el producto en el catálogo.")

    def _abrir_buscador(self):
        from ui.pos import BuscadorProductos
        dlg = BuscadorProductos(self)
        dlg.producto_seleccionado.connect(self._agregar_producto)
        dlg.exec()
        QTimer.singleShot(100, self.scan_input.setFocus)

    def _agregar_producto(self, producto: dict):
        pid = producto["id"]
        upc = producto.get("unidades_por_caja") or 1
        if pid in self.items:
            self.items[pid]["cantidad"] += 1
        else:
            modo_default = "caja" if upc > 1 else "unidad"
            self.items[pid] = {
                "producto": producto, "cantidad": 1, "modo": modo_default,
                "lote": {"cosecha": None, "fecha_vencimiento": None, "notas": ""}
            }
        self._refrescar_tabla()
        QTimer.singleShot(50, self.scan_input.setFocus)

    def _total_unidades(self, entry: dict) -> int:
        """Calcula las unidades reales a agregar según modo."""
        upc = entry["producto"].get("unidades_por_caja") or 1
        if entry["modo"] == "caja" and upc > 1:
            return entry["cantidad"] * upc
        return entry["cantidad"]

    def _refrescar_tabla(self):
        items_list = list(self.items.values())
        self.tabla.setRowCount(len(items_list))
        total_unidades = 0

        for i, entry in enumerate(items_list):
            p = entry["producto"]
            pid = p["id"]
            upc = p.get("unidades_por_caja") or 1

            self.tabla.setItem(i, 0, QTableWidgetItem(p.get("codigo_barras") or "S/C"))
            self.tabla.setItem(i, 1, QTableWidgetItem(p["nombre"]))
            self.tabla.setItem(i, 2, QTableWidgetItem(str(p["stock_actual"])))

            # Col 3: Modo de ingreso
            if upc > 1:
                cmb_modo = QComboBox()
                cmb_modo.addItem(f"📦 Por caja  ({upc} unid.)", "caja")
                cmb_modo.addItem("🍷 Por unidad", "unidad")
                if entry["modo"] == "unidad":
                    cmb_modo.setCurrentIndex(1)
                cmb_modo.setStyleSheet(
                    "background:#2C2C2C; color:#F5F5F5; border:1px solid #444; padding:4px;"
                )
                cmb_modo.currentIndexChanged.connect(
                    lambda _, pid=pid, row=i: self._cambiar_modo(pid, row)
                )
                self.tabla.setCellWidget(i, 3, cmb_modo)
            else:
                lbl_u = QLabel("🍷 Por unidad")
                lbl_u.setStyleSheet("color:#888; padding:4px 8px;")
                self.tabla.setCellWidget(i, 3, lbl_u)

            # Col 4: Cantidad
            spin = QSpinBox()
            spin.setRange(1, 99999)
            spin.setValue(entry["cantidad"])
            spin.setStyleSheet("background:#2C2C2C; color:#F5F5F5; border:1px solid #444;")
            spin.valueChanged.connect(lambda v, pid=pid, row=i: self._actualizar_cantidad(pid, v, row))
            self.tabla.setCellWidget(i, 4, spin)

            # Col 5: Total unidades
            total = self._total_unidades(entry)
            total_unidades += total
            lbl_total = QLabel(f"<b>{total}</b> unidades")
            lbl_total.setAlignment(Qt.AlignmentFlag.AlignCenter)
            color = "#C9A84C" if (upc > 1 and entry["modo"] == "caja") else "#AAAAAA"
            lbl_total.setStyleSheet(f"color:{color}; font-size:11pt; padding:4px;")
            self.tabla.setCellWidget(i, 5, lbl_total)

            # Col 6: Año de cosecha (spinner inline)
            lote = entry.get("lote", {})
            spin_c = QSpinBox()
            spin_c.setRange(0, 2100)
            spin_c.setSpecialValueText("—")
            spin_c.setValue(lote.get("cosecha") or 0)
            spin_c.setToolTip("Año de cosecha / producción")
            spin_c.setStyleSheet("background:#2C2C2C; color:#C9A84C; border:1px solid #444;")
            spin_c.valueChanged.connect(
                lambda v, p=pid: self._set_lote_field(p, "cosecha", v if v > 0 else None))
            self.tabla.setCellWidget(i, 6, spin_c)

            # Col 7: Fecha de vencimiento (date picker + botón limpiar)
            min_date = QDate(2000, 1, 1)
            date_v = QDateEdit()
            date_v.setCalendarPopup(True)
            date_v.setDisplayFormat("dd/MM/yy")
            date_v.setSpecialValueText("— S/F")
            date_v.setMinimumDate(min_date)
            date_v.setDate(
                QDate.fromString(lote["fecha_vencimiento"], "yyyy-MM-dd")
                if lote.get("fecha_vencimiento")
                else min_date
            )
            date_v.setStyleSheet("background:#2C2C2C; color:#AAAAAA; border:1px solid #444;")
            date_v.dateChanged.connect(
                lambda d, p=pid: self._set_lote_field(
                    p, "fecha_vencimiento",
                    d.toString("yyyy-MM-dd") if d != min_date else None))

            btn_clear_v = QPushButton("✕")
            btn_clear_v.setFixedSize(22, 22)
            btn_clear_v.setToolTip("Quitar fecha de vencimiento")
            btn_clear_v.setStyleSheet(
                "QPushButton{background:#5C0000;color:#FF8A80;border-radius:4px;"
                "font-size:9pt;font-weight:700;border:none;}"
                "QPushButton:hover{background:#7F0000;}")
            btn_clear_v.clicked.connect(lambda _, dv=date_v, md=min_date: dv.setDate(md))

            cell_venc = QWidget()
            cell_lay  = QHBoxLayout(cell_venc)
            cell_lay.setContentsMargins(2, 2, 2, 2)
            cell_lay.setSpacing(3)
            cell_lay.addWidget(date_v)
            cell_lay.addWidget(btn_clear_v)
            self.tabla.setCellWidget(i, 7, cell_venc)
            self.tabla.setRowHeight(i, 56)

        # Resumen total
        n = len(items_list)
        self.lbl_resumen.setText(
            f"📊  {n} producto(s)  →  {total_unidades} unidades en total a ingresar" if n else ""
        )

    def _set_lote_field(self, pid: int, campo: str, valor):
        if pid in self.items:
            self.items[pid]["lote"][campo] = valor

    def _cambiar_modo(self, pid: int, row: int):
        if pid not in self.items:
            return
        cmb = self.tabla.cellWidget(row, 3)
        if cmb and hasattr(cmb, 'currentData'):
            self.items[pid]["modo"] = cmb.currentData()
        self._actualizar_fila_total(pid, row)
        self._actualizar_resumen_label()

    def _actualizar_cantidad(self, pid: int, valor: int, row: int):
        if pid in self.items:
            self.items[pid]["cantidad"] = valor
            self._actualizar_fila_total(pid, row)
            self._actualizar_resumen_label()

    def _actualizar_fila_total(self, pid: int, row: int):
        """Actualiza solo la celda de total unidades para no reconstruir toda la tabla."""
        if pid not in self.items:
            return
        entry = self.items[pid]
        total = self._total_unidades(entry)
        upc = entry["producto"].get("unidades_por_caja") or 1
        lbl = QLabel(f"<b>{total}</b> unidades")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        color = "#C9A84C" if (upc > 1 and entry["modo"] == "caja") else "#AAAAAA"
        lbl.setStyleSheet(f"color:{color}; font-size:11pt; padding:4px;")
        self.tabla.setCellWidget(row, 5, lbl)

    def _actualizar_resumen_label(self):
        total_u = sum(self._total_unidades(e) for e in self.items.values())
        n = len(self.items)
        self.lbl_resumen.setText(
            f"📊  {n} producto(s)  →  {total_u} unidades en total a ingresar" if n else ""
        )

    def _confirmar(self):
        if not self.items:
            QMessageBox.information(self, "Vacío", "No hay productos para ingresar.")
            return
        motivo = self.txt_motivo.text().strip() or "Ingreso de mercadería"
        total_unidades = sum(self._total_unidades(e) for e in self.items.values())
        try:
            for pid, entry in self.items.items():
                unidades_a_agregar = self._total_unidades(entry)
                upc = entry["producto"].get("unidades_por_caja") or 1
                if entry["modo"] == "caja" and upc > 1:
                    detalle_motivo = f"{motivo} ({entry['cantidad']} caja(s) × {upc} unid.)"
                else:
                    detalle_motivo = motivo
                db.agregar_stock(pid, unidades_a_agregar, detalle_motivo)
                # Siempre registrar lote para mantener trazabilidad exacta
                lote = entry.get("lote", {})
                db.crear_lote(
                    pid, unidades_a_agregar,
                    fecha_vencimiento=lote.get("fecha_vencimiento"),
                    cosecha=lote.get("cosecha"),
                    motivo=detalle_motivo,
                    notas=lote.get("notas", "")
                )
            QMessageBox.information(
                self, "✅  Stock actualizado",
                f"Se ingresaron {total_unidades} unidades en {len(self.items)} producto(s)."
            )
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


# ─────────────────────────────────────────────────────────────
#  Widget principal de Stock
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
#  Diálogo: Pedido de compra / Restockeo
# ─────────────────────────────────────────────────────────────

class DialogoPedidoCompra(QDialog):
    """
    Muestra todos los productos con bajo stock con un spinner por fila.
    El usuario ajusta cuánto comprar de cada uno y genera un Excel
    formateado listo para enviar al proveedor.
    """

    def __init__(self, productos: list, parent=None):
        super().__init__(parent)
        self.productos = productos
        self.spinners  = {}          # producto_id → QSpinBox
        self.setWindowTitle("📋  Pedido de compra")
        self.setMinimumSize(880, 520)
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 20, 20, 20)

        titulo = QLabel("📋  Pedido de compra")
        titulo.setObjectName("titulo_seccion")
        lay.addWidget(titulo)

        info = QLabel(
            "Ajustá la cantidad a comprar para cada producto. "
            "Ponelo en 0 para excluirlo del pedido."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#999;font-size:12px;")
        lay.addWidget(info)

        # ── Tabla ────────────────────────────────────────────
        cols = ["Proveedor", "Producto", "Categoría",
                "Stock\nactual", "Mínimo", "Faltante",
                "Unid/\ncaja", "📦 Cantidad\na comprar"]
        self.tabla = QTableWidget(len(self.productos), len(cols))
        self.tabla.setHorizontalHeaderLabels(cols)
        self.tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tabla.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        for c in (3, 4, 5, 6, 7):
            self.tabla.horizontalHeader().setSectionResizeMode(
                c, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        for i, p in enumerate(self.productos):
            faltante = max(0, p["stock_minimo"] - p["stock_actual"])
            upc      = p["unidades_por_caja"] or 1
            proveedor = p["proveedor_nombre"] or "Sin proveedor"

            self.tabla.setItem(i, 0, QTableWidgetItem(proveedor))
            self.tabla.setItem(i, 1, QTableWidgetItem(p["nombre"]))
            self.tabla.setItem(i, 2, QTableWidgetItem(p["categoria_nombre"] or ""))

            def _centered(text, color=None):
                it = QTableWidgetItem(text)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if color:
                    it.setForeground(QColor(color))
                return it

            stock_color = "#F44336" if p["stock_actual"] == 0 else "#FF9800"
            self.tabla.setItem(i, 3, _centered(str(p["stock_actual"]), stock_color))
            self.tabla.setItem(i, 4, _centered(str(p["stock_minimo"])))
            self.tabla.setItem(i, 5, _centered(str(faltante), "#C9A84C"))
            self.tabla.setItem(i, 6, _centered(str(upc) if upc > 1 else "—"))

            spin = QSpinBox()
            spin.setRange(0, 99999)
            spin.setValue(faltante)
            spin.setAlignment(Qt.AlignmentFlag.AlignCenter)
            spin.setStyleSheet(
                "QSpinBox{background:#1A1A1A;color:#C9A84C;font-weight:700;"
                "border:1px solid #C9A84C;border-radius:4px;padding:2px 6px;}"
                "QSpinBox::up-button,QSpinBox::down-button{width:18px;}"
            )
            self.spinners[p["id"]] = spin
            self.tabla.setCellWidget(i, 7, spin)
            self.tabla.setRowHeight(i, 44)

        lay.addWidget(self.tabla, 1)

        # ── Resumen ──────────────────────────────────────────
        self.lbl_total = QLabel()
        self.lbl_total.setStyleSheet("color:#FF9800;font-size:12px;")
        lay.addWidget(self.lbl_total)
        self._actualizar_lbl_total()

        # ── Botones ──────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)

        btn_excel = QPushButton("📊  Generar Excel")
        btn_excel.setObjectName("btn_exito")
        btn_excel.setMinimumWidth(180)
        btn_excel.clicked.connect(self._generar_excel)
        btn_row.addWidget(btn_excel)

        lay.addLayout(btn_row)

    def _actualizar_lbl_total(self):
        total = sum(
            self.spinners[p["id"]].value()
            for p in self.productos
            if self.spinners[p["id"]].value() > 0
        )
        incl = sum(1 for p in self.productos if self.spinners[p["id"]].value() > 0)
        self.lbl_total.setText(
            f"⚠️  {len(self.productos)} producto(s) con bajo stock  •  "
            f"📦 {incl} incluidos en el pedido  •  Total unidades: {total}"
        )

    def _generar_excel(self):
        from PyQt6.QtWidgets import QFileDialog
        import pandas as pd
        from datetime import date as _date
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        filas = []
        for p in self.productos:
            cant = self.spinners[p["id"]].value()
            if cant == 0:
                continue
            filas.append({
                "Proveedor":         p["proveedor_nombre"] or "Sin proveedor",
                "Teléfono":          p["proveedor_telefono"] or "",
                "Email":             p["proveedor_email"] or "",
                "Código de barras":  p["codigo_barras"] or "",
                "Producto":          p["nombre"],
                "Categoría":         p["categoria_nombre"] or "",
                "Stock actual":      p["stock_actual"],
                "Stock mínimo":      p["stock_minimo"],
                "Faltante":          max(0, p["stock_minimo"] - p["stock_actual"]),
                "Cantidad a pedir":  cant,
                "Precio costo ref.": round(p["precio_costo"] or 0, 2),
                "Subtotal ref.":     round((p["precio_costo"] or 0) * cant, 2),
            })

        if not filas:
            QMessageBox.warning(
                self, "Sin productos",
                "Todas las cantidades están en 0.\nAjustá antes de exportar.")
            return

        fecha_str   = _date.today().strftime("%Y-%m-%d")
        fecha_larga = _date.today().strftime("%d/%m/%Y")
        nombre_def  = f"Pedido_Compra_{fecha_str}.xlsx"

        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar pedido de compra", nombre_def, "Excel (*.xlsx)")
        if not ruta:
            return

        df = pd.DataFrame(filas)

        # ─── Construir el Excel con openpyxl ─────────────────
        THIN = Border(
            left=Side(style="thin",   color="555555"),
            right=Side(style="thin",  color="555555"),
            top=Side(style="thin",    color="555555"),
            bottom=Side(style="thin", color="555555"),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedido de Compra"

        ncols = len(df.columns)

        # ── Fila de título ────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ct = ws.cell(row=1, column=1)
        ct.value = f"PEDIDO DE COMPRA  —  {fecha_larga}"
        ct.font  = Font(bold=True, size=14, color="C9A84C")
        ct.fill  = PatternFill("solid", fgColor="1A1A1A")
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # ── Fila de headers ───────────────────────────────────
        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="722F37")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN
        ws.row_dimensions[2].height = 32

        # ── Filas de datos, agrupadas por proveedor ───────────
        idx_pedir   = headers.index("Cantidad a pedir")
        idx_costo   = headers.index("Precio costo ref.")
        idx_subtotal = headers.index("Subtotal ref.")
        idx_texto   = {0, 1, 2, 3, 4, 5}   # columnas de texto (izquierda)

        current_prov = None
        current_row  = 3

        for _, row in df.iterrows():
            prov = row["Proveedor"]

            # ── Separador de proveedor ────────────────────────
            if prov != current_prov:
                if current_prov is not None:
                    current_row += 1   # fila en blanco entre proveedores

                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=ncols)
                cs = ws.cell(row=current_row, column=1)
                cs.value     = f"  ▶  {prov}"
                cs.font      = Font(bold=True, color="C9A84C", size=11)
                cs.fill      = PatternFill("solid", fgColor="3D1A1E")
                cs.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                current_prov = prov

            # ── Fila de producto ──────────────────────────────
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.border    = THIN
                cell.alignment = Alignment(
                    horizontal="left" if (ci - 1) in idx_texto else "center",
                    vertical="center"
                )
                if (ci - 1) == idx_pedir:
                    cell.font = Font(bold=True, color="C9A84C", size=11)
                    cell.fill = PatternFill("solid", fgColor="2A1800")
                if (ci - 1) in (idx_costo, idx_subtotal):
                    cell.number_format = '"$"#,##0.00'
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # ── Fila TOTAL ────────────────────────────────────────
        current_row += 1
        total_vals = [""] * ncols
        total_vals[0]          = "TOTAL"
        total_vals[idx_pedir]  = int(df["Cantidad a pedir"].sum())
        total_vals[idx_subtotal] = round(df["Subtotal ref."].sum(), 2)

        for ci, val in enumerate(total_vals, 1):
            cell = ws.cell(row=current_row, column=ci, value=val)
            cell.font      = Font(bold=True, color="C9A84C", size=11)
            cell.fill      = PatternFill("solid", fgColor="1A1A1A")
            cell.border    = THIN
            cell.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center"
            )
        ws.cell(row=current_row,
                column=idx_subtotal + 1).number_format = '"$"#,##0.00'
        ws.row_dimensions[current_row].height = 26

        # ── Anchos de columna ─────────────────────────────────
        anchos = [22, 13, 26, 18, 28, 16, 11, 11, 11, 16, 15, 14]
        for i, w in enumerate(anchos[:ncols], 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = "A3"

        wb.save(ruta)
        QMessageBox.information(
            self, "✅  Pedido exportado",
            f"Guardado en:\n{ruta}\n\n"
            f"📦  {len(filas)} producto(s) incluidos en el pedido."
        )
        self.accept()


class StockWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.cargar_productos()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(16, 16, 16, 16)

        # Encabezado
        header = QHBoxLayout()
        titulo = QLabel("📦  Stock y Productos")
        titulo.setObjectName("titulo_seccion")
        header.addWidget(titulo)
        header.addStretch()

        btn_nuevo = QPushButton("➕  Nuevo producto")
        btn_nuevo.clicked.connect(self._nuevo_producto)
        header.addWidget(btn_nuevo)

        btn_carga = QPushButton("📥  Cargar stock")
        btn_carga.setObjectName("btn_exito")
        btn_carga.clicked.connect(self._cargar_stock)
        header.addWidget(btn_carga)

        lay.addLayout(header)

        # Alertas de stock bajo
        self.lbl_alerta = QLabel("")
        self.lbl_alerta.setObjectName("alerta_stock")
        self.lbl_alerta.setWordWrap(True)
        lay.addWidget(self.lbl_alerta)

        # Tabs: Catálogo | Sin rotación | Historial
        tabs = QTabWidget()
        tabs.setDocumentMode(True)

        # Tab 1: Catálogo
        tab_catalogo = QWidget()
        lay_cat = QVBoxLayout(tab_catalogo)
        lay_cat.setSpacing(8)

        # Filtros
        filtro_row = QHBoxLayout()
        self.txt_filtro = QLineEdit()
        self.txt_filtro.setPlaceholderText("🔍  Filtrar por nombre o código…")
        self.txt_filtro.textChanged.connect(self._filtrar)
        filtro_row.addWidget(self.txt_filtro, 1)

        self.cmb_cat_filtro = QComboBox()
        self.cmb_cat_filtro.addItem("Todas las categorías", None)
        for c in db.obtener_categorias():
            self.cmb_cat_filtro.addItem(c["nombre"], c["id"])
        self.cmb_cat_filtro.currentIndexChanged.connect(self._filtrar)
        filtro_row.addWidget(self.cmb_cat_filtro)
        lay_cat.addLayout(filtro_row)

        self.tabla_productos = QTableWidget(0, 8)
        self.tabla_productos.setHorizontalHeaderLabels([
            "ID", "Código", "Nombre", "Categoría",
            "Precio venta", "Costo", "Stock", "Acciones"
        ])
        self.tabla_productos.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.tabla_productos.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla_productos.setColumnWidth(0, 50)
        self.tabla_productos.setColumnWidth(1, 120)
        self.tabla_productos.setColumnWidth(4, 110)
        self.tabla_productos.setColumnWidth(5, 90)
        self.tabla_productos.setColumnWidth(6, 80)
        self.tabla_productos.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.tabla_productos.setColumnWidth(7, 215)
        self.tabla_productos.setAlternatingRowColors(False)
        self.tabla_productos.setItemDelegate(FilaColorDelegate(self.tabla_productos))
        self.tabla_productos.verticalHeader().setVisible(False)
        self.tabla_productos.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tabla_productos.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay_cat.addWidget(self.tabla_productos)

        tab_catalogo.setLayout(lay_cat)
        tabs.addTab(tab_catalogo, "📋  Catálogo")

        # Tab 1: Precios
        tab_precios = self._build_tab_precios()
        tabs.addTab(tab_precios, "💰  Precios")

        # Tab 2: Restockear
        tab_restock = QWidget()
        lay_rs = QVBoxLayout(tab_restock)
        lay_rs.setSpacing(8)

        rs_header = QHBoxLayout()
        lbl_rs = QLabel("Productos con stock por debajo del mínimo configurado:")
        lbl_rs.setStyleSheet("color:#FF9800;")
        rs_header.addWidget(lbl_rs, 1)
        btn_refresh_rs = QPushButton("🔄  Actualizar")
        btn_refresh_rs.clicked.connect(self._cargar_restock)
        rs_header.addWidget(btn_refresh_rs)
        btn_exportar_pedido = QPushButton("📊  Exportar pedido a Excel")
        btn_exportar_pedido.setObjectName("btn_exito")
        btn_exportar_pedido.clicked.connect(self._exportar_pedido)
        rs_header.addWidget(btn_exportar_pedido)
        lay_rs.addLayout(rs_header)

        self.tabla_restock = QTableWidget(0, 7)
        self.tabla_restock.setHorizontalHeaderLabels([
            "Proveedor", "Producto", "Categoría",
            "Stock actual", "Mínimo", "Faltante", "Unid/caja"
        ])
        self.tabla_restock.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.tabla_restock.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla_restock.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla_restock.setAlternatingRowColors(True)
        self.tabla_restock.verticalHeader().setVisible(False)
        self.tabla_restock.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tabla_restock.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        lay_rs.addWidget(self.tabla_restock, 1)

        tab_restock.setLayout(lay_rs)
        tabs.addTab(tab_restock, "🛒  Restockear")

        # Tab 2: Próximos vencimientos
        tab_venc = QWidget()
        lay_venc = QVBoxLayout(tab_venc)
        lay_venc.setSpacing(8)

        venc_header = QHBoxLayout()
        venc_header.addWidget(QLabel("Mostrar vencimientos en los próximos:"))
        self.spin_dias_venc = QSpinBox()
        self.spin_dias_venc.setRange(1, 365)
        self.spin_dias_venc.setValue(30)
        self.spin_dias_venc.setSuffix(" días")
        venc_header.addWidget(self.spin_dias_venc)
        btn_refresh_venc = QPushButton("🔄  Actualizar")
        btn_refresh_venc.clicked.connect(self._cargar_vencimientos)
        venc_header.addWidget(btn_refresh_venc)
        venc_header.addStretch()
        lay_venc.addLayout(venc_header)

        self.tabla_vencimientos = QTableWidget(0, 6)
        self.tabla_vencimientos.setHorizontalHeaderLabels([
            "Producto", "Cosecha", "Cant. lote", "Vencimiento", "Días restantes", "Motivo"
        ])
        self.tabla_vencimientos.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.tabla_vencimientos.setColumnWidth(1, 80)
        self.tabla_vencimientos.setColumnWidth(2, 90)
        self.tabla_vencimientos.setColumnWidth(3, 110)
        self.tabla_vencimientos.setColumnWidth(4, 120)
        self.tabla_vencimientos.setAlternatingRowColors(True)
        self.tabla_vencimientos.verticalHeader().setVisible(False)
        self.tabla_vencimientos.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay_venc.addWidget(self.tabla_vencimientos)

        tab_venc.setLayout(lay_venc)
        tabs.addTab(tab_venc, "⏰  Vencimientos")

        # Tab 3: Sin rotación
        tab_sin_rot = QWidget()
        lay_sr = QVBoxLayout(tab_sin_rot)

        sr_header = QHBoxLayout()
        self.spin_dias_sin_rot = QSpinBox()
        self.spin_dias_sin_rot.setRange(7, 365)
        self.spin_dias_sin_rot.setValue(90)
        self.spin_dias_sin_rot.setSuffix(" días sin venta")
        sr_header.addWidget(QLabel("Mostrar productos con más de:"))
        sr_header.addWidget(self.spin_dias_sin_rot)
        btn_refresh_sr = QPushButton("Actualizar")
        btn_refresh_sr.clicked.connect(self._cargar_sin_rotacion)
        sr_header.addWidget(btn_refresh_sr)
        sr_header.addStretch()
        lay_sr.addLayout(sr_header)

        self.tabla_sin_rotacion = QTableWidget(0, 5)
        self.tabla_sin_rotacion.setHorizontalHeaderLabels([
            "Producto", "Categoría", "Stock", "Última venta", "Días sin rotar"
        ])
        self.tabla_sin_rotacion.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tabla_sin_rotacion.setAlternatingRowColors(True)
        self.tabla_sin_rotacion.verticalHeader().setVisible(False)
        lay_sr.addWidget(self.tabla_sin_rotacion)
        tab_sin_rot.setLayout(lay_sr)
        tabs.addTab(tab_sin_rot, "🕰️  Sin Rotación")

        # Tab 3: Historial movimientos
        tab_historial = QWidget()
        lay_h = QVBoxLayout(tab_historial)

        self.tabla_historial = QTableWidget(0, 6)
        self.tabla_historial.setHorizontalHeaderLabels([
            "Fecha", "Producto", "Tipo", "Cantidad", "Stock antes", "Stock después"
        ])
        self.tabla_historial.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tabla_historial.setAlternatingRowColors(True)
        self.tabla_historial.verticalHeader().setVisible(False)
        lay_h.addWidget(self.tabla_historial)
        tab_historial.setLayout(lay_h)
        tabs.addTab(tab_historial, "📜  Historial")

        tabs.currentChanged.connect(self._tab_changed)
        lay.addWidget(tabs, 1)
        self.tabs = tabs

    def cargar_productos(self):
        self.todos_productos = db.obtener_todos_productos()
        # Pre-cargar cosechas y alertas de vencimiento de todos los productos
        self._cosechas_cache = {}
        self._vencimiento_cache = set()   # ids de productos con lote próximo a vencer
        dias_alerta = int(db.get_config("dias_alerta_vencimiento", 7))
        lotes_venciendo = db.lotes_por_vencer(dias_alerta)
        for lv in lotes_venciendo:
            self._vencimiento_cache.add(lv["producto_id"])
        for p in self.todos_productos:
            lotes = db.obtener_lotes_producto(p["id"])
            cosechas = sorted({l["cosecha"] for l in lotes if l["cosecha"]})
            if cosechas:
                self._cosechas_cache[p["id"]] = cosechas
        self._mostrar_productos(self.todos_productos)
        self._mostrar_alertas()

    def _mostrar_alertas(self):
        msgs = []
        bajos = db.productos_bajo_stock()
        if bajos:
            nombres = ", ".join(p["nombre"] for p in bajos[:5])
            extra   = f" +{len(bajos)-5} más" if len(bajos) > 5 else ""
            msgs.append(f"⚠️  Stock bajo: {nombres}{extra}")
        venciendo = db.lotes_por_vencer(int(db.get_config("dias_alerta_vencimiento", 7)))
        if venciendo:
            nom_v = ", ".join(l["producto_nombre"] for l in venciendo[:3])
            extra_v = f" +{len(venciendo)-3} más" if len(venciendo) > 3 else ""
            msgs.append(f"⏰  Próx. vencim.: {nom_v}{extra_v}")
        self.lbl_alerta.setText("   |   ".join(msgs) if msgs else "")

    def _mostrar_productos(self, productos: list):
        self.tabla_productos.setRowCount(len(productos))
        for i, p in enumerate(productos):
            self.tabla_productos.setItem(i, 0, QTableWidgetItem(str(p["id"])))
            self.tabla_productos.setItem(i, 1, QTableWidgetItem(
                p["codigo_barras"] or ("S/C" if p["sin_codigo"] else "—")))

            # Nombre + cosechas inline
            nombre    = p["nombre"]
            cosechas  = getattr(self, "_cosechas_cache", {}).get(p["id"], [])
            if cosechas:
                años_txt = "  🍷 " + " · ".join(str(c) for c in cosechas)
                nombre_item = QTableWidgetItem(nombre + años_txt)
                nombre_item.setToolTip("Cosechas en stock: " + ", ".join(str(c) for c in cosechas))
                nombre_item.setForeground(QColor("#C9A84C"))
            else:
                nombre_item = QTableWidgetItem(nombre)
            self.tabla_productos.setItem(i, 2, nombre_item)
            self.tabla_productos.setItem(i, 3, QTableWidgetItem(p["categoria_nombre"] or ""))
            self.tabla_productos.setItem(i, 4, QTableWidgetItem(f"${p['precio_venta']:,.2f}"))
            self.tabla_productos.setItem(i, 5, QTableWidgetItem(f"${p['precio_costo']:,.2f}"))

            stock_item = QTableWidgetItem(str(p["stock_actual"]))
            stock_bajo = p["stock_actual"] <= p["stock_minimo"]
            vence_pronto = p["id"] in getattr(self, "_vencimiento_cache", set())
            if stock_bajo:
                stock_item.setForeground(QColor("#FF9800"))
                stock_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            self.tabla_productos.setItem(i, 6, stock_item)

            # Colorear fila completa según estado
            if stock_bajo:
                row_bg = QColor(110, 20, 22)          # rojo oscuro opaco
            elif vence_pronto:
                row_bg = QColor(100, 55, 0)           # naranja oscuro opaco
            elif i % 2 == 1:
                row_bg = QColor(37, 37, 37)           # alternado normal
            else:
                row_bg = QColor(30, 30, 30)           # normal

            for col in range(7):
                it = self.tabla_productos.item(i, col)
                if it:
                    it.setBackground(row_bg)

            # Botones de acciones
            acc_widget = QWidget()
            acc_lay = QHBoxLayout(acc_widget)
            acc_lay.setContentsMargins(8, 4, 8, 4)
            acc_lay.setSpacing(8)

            producto_dict = dict(p)

            # Botón Opciones — menú de acciones
            btn_acc = QPushButton("☰  Opciones")
            btn_acc.setFixedHeight(32)
            btn_acc.setMinimumWidth(95)
            btn_acc.setToolTip("Editar • Agregar stock • Ajuste • Ver lotes")
            btn_acc.setStyleSheet(
                "QPushButton{background:#2C2C2C;border:1px solid #555;border-radius:6px;"
                "color:#F5F5F5;font-size:10pt;padding:0 12px;}"
                "QPushButton:hover{background:#3C3C3C;border-color:#888;}"
            )

            def _make_menu(pd=producto_dict):
                menu = QMenu(self)
                menu.setStyleSheet(
                    "QMenu{background:#252525;border:1px solid #444;border-radius:6px;}"
                    "QMenu::item{padding:9px 22px;color:#F5F5F5;font-size:11pt;}"
                    "QMenu::item:selected{background:#722F37;border-radius:4px;}"
                )
                menu.addAction("✏️  Editar producto",
                               lambda: self._editar_producto(pd))
                menu.addAction("➕  Agregar stock",
                               lambda: self._agregar_stock_rapido(pd))
                menu.addAction("⚖️  Ajuste de stock",
                               lambda: self._ajustar_stock_producto(pd))
                menu.addAction("📋  Ver lotes / cosechas",
                               lambda: self._ver_lotes(pd))
                return menu

            btn_acc.clicked.connect(
                lambda _, b=btn_acc, pd=producto_dict:
                    _make_menu(pd).exec(b.mapToGlobal(
                        b.rect().bottomLeft())))

            # Botón eliminar
            btn_del = QPushButton("🗑  Borrar")
            btn_del.setFixedHeight(32)
            btn_del.setMinimumWidth(85)
            btn_del.setToolTip("Eliminar producto")
            btn_del.setStyleSheet(
                "QPushButton{background:#7F0000;color:white;border-radius:6px;"
                "font-size:10pt;padding:0 12px;}"
                "QPushButton:hover{background:#B71C1C;}"
            )
            btn_del.clicked.connect(
                lambda _, pd=producto_dict: self._eliminar_producto(pd))

            acc_lay.addWidget(btn_acc)
            acc_lay.addWidget(btn_del)
            self.tabla_productos.setCellWidget(i, 7, acc_widget)
            self.tabla_productos.setRowHeight(i, 48)

    def _filtrar(self):
        texto = self.txt_filtro.text().lower()
        cat_id = self.cmb_cat_filtro.currentData()
        resultado = [
            p for p in self.todos_productos
            if (texto in p["nombre"].lower() or
                texto in (p["codigo_barras"] or "").lower())
            and (cat_id is None or p["categoria_id"] == cat_id)
        ]
        self._mostrar_productos(resultado)

    def _tab_changed(self, idx):
        if idx == 1:
            self._prec_cargar()
        elif idx == 2:
            self._cargar_restock()
        elif idx == 3:
            self._cargar_vencimientos()
        elif idx == 4:
            self._cargar_sin_rotacion()
        elif idx == 5:
            self._cargar_historial()

    def _cargar_restock(self):
        productos = db.productos_para_restock()
        self.tabla_restock.setRowCount(len(productos))
        for i, p in enumerate(productos):
            faltante = max(0, p["stock_minimo"] - p["stock_actual"])
            upc      = p["unidades_por_caja"] or 1

            self.tabla_restock.setItem(i, 0, QTableWidgetItem(
                p["proveedor_nombre"] or "Sin proveedor"))
            self.tabla_restock.setItem(i, 1, QTableWidgetItem(p["nombre"]))
            self.tabla_restock.setItem(i, 2, QTableWidgetItem(
                p["categoria_nombre"] or ""))

            def _c(val, color=None):
                it = QTableWidgetItem(str(val))
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if color:
                    it.setForeground(QColor(color))
                return it

            stock_color = "#F44336" if p["stock_actual"] == 0 else "#FF9800"
            self.tabla_restock.setItem(i, 3, _c(p["stock_actual"], stock_color))
            self.tabla_restock.setItem(i, 4, _c(p["stock_minimo"]))
            self.tabla_restock.setItem(i, 5, _c(faltante, "#C9A84C"))
            self.tabla_restock.setItem(i, 6,
                _c(str(upc) if upc > 1 else "—"))
            self.tabla_restock.setRowHeight(i, 38)

        if len(productos) == 0:
            self.tabla_restock.setRowCount(1)
            item = QTableWidgetItem("✅  Todos los productos tienen stock suficiente")
            item.setForeground(QColor("#4CAF50"))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla_restock.setItem(0, 0, item)
            self.tabla_restock.setSpan(0, 0, 1, 7)

    def _exportar_pedido(self):
        productos = db.productos_para_restock()
        if not productos:
            QMessageBox.information(
                self, "Sin faltantes",
                "✅  Todos los productos tienen stock suficiente.\n"
                "No hay nada que pedir por ahora.")
            return
        dlg = DialogoPedidoCompra(productos, self)
        dlg.exec()

    def _cargar_sin_rotacion(self):
        dias = self.spin_dias_sin_rot.value()
        productos = db.productos_sin_rotacion(dias)
        self.tabla_sin_rotacion.setRowCount(len(productos))
        from datetime import date
        hoy = date.today()
        for i, p in enumerate(productos):
            self.tabla_sin_rotacion.setItem(i, 0, QTableWidgetItem(p["nombre"]))
            self.tabla_sin_rotacion.setItem(i, 2, QTableWidgetItem(str(p["stock_actual"])))
            ultima = p["ultima_venta_fecha"] or "Nunca"
            self.tabla_sin_rotacion.setItem(i, 3, QTableWidgetItem(ultima))
            if ultima != "Nunca":
                from datetime import date as d
                dias_sin = (hoy - d.fromisoformat(ultima)).days
                self.tabla_sin_rotacion.setItem(i, 4, QTableWidgetItem(str(dias_sin)))
            else:
                self.tabla_sin_rotacion.setItem(i, 4, QTableWidgetItem("∞"))

    def _cargar_historial(self):
        movs = db.historial_movimientos(limite=200)
        self.tabla_historial.setRowCount(len(movs))
        tipo_colores = {
            "entrada": "#2E7D32", "salida": "#B71C1C",
            "ajuste": "#E65100", "devolucion": "#0D47A1"
        }
        for i, m in enumerate(movs):
            fecha_str = str(m["fecha"])[:16]
            self.tabla_historial.setItem(i, 0, QTableWidgetItem(fecha_str))
            self.tabla_historial.setItem(i, 1, QTableWidgetItem(m["producto_nombre"]))
            tipo_item = QTableWidgetItem(m["tipo"].capitalize())
            color = tipo_colores.get(m["tipo"], "#FFFFFF")
            tipo_item.setForeground(QColor(color))
            self.tabla_historial.setItem(i, 2, tipo_item)
            self.tabla_historial.setItem(i, 3, QTableWidgetItem(str(m["cantidad"])))
            self.tabla_historial.setItem(i, 4, QTableWidgetItem(
                str(m["stock_anterior"]) if m["stock_anterior"] is not None else "—"))
            self.tabla_historial.setItem(i, 5, QTableWidgetItem(
                str(m["stock_nuevo"]) if m["stock_nuevo"] is not None else "—"))

    def _nuevo_producto(self):
        dlg = DialogoProducto(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.cargar_productos()

    def _editar_producto(self, producto: dict):
        dlg = DialogoProducto(self, producto=producto)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.cargar_productos()

    # ── Tab Precios ───────────────────────────────────────────

    def _build_tab_precios(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.setContentsMargins(12, 12, 12, 12)

        # ── Filtros ──────────────────────────────────────────
        filtro_row = QHBoxLayout()
        self.prec_txt_filtro = QLineEdit()
        self.prec_txt_filtro.setPlaceholderText("🔍  Buscar producto por nombre…")
        self.prec_txt_filtro.textChanged.connect(self._prec_filtrar)
        filtro_row.addWidget(self.prec_txt_filtro, 1)

        self.prec_cmb_cat = QComboBox()
        self.prec_cmb_cat.addItem("Todas las categorías", None)
        for c in db.obtener_categorias():
            self.prec_cmb_cat.addItem(c["nombre"], c["id"])
        self.prec_cmb_cat.currentIndexChanged.connect(self._prec_filtrar)
        filtro_row.addWidget(self.prec_cmb_cat)
        lay.addLayout(filtro_row)

        # ── Tabla ─────────────────────────────────────────────
        self.prec_tabla = QTableWidget(0, 6)
        self.prec_tabla.setHorizontalHeaderLabels([
            "☑", "Producto", "Categoría", "Precio actual", "Nuevo precio", "Cambio"
        ])
        self.prec_tabla.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.prec_tabla.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self.prec_tabla.setColumnWidth(0, 44)
        self.prec_tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.prec_tabla.setColumnWidth(3, 120)
        self.prec_tabla.setColumnWidth(4, 140)
        self.prec_tabla.setColumnWidth(5, 90)
        self.prec_tabla.setAlternatingRowColors(True)
        self.prec_tabla.verticalHeader().setVisible(False)
        self.prec_tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.prec_tabla.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        lay.addWidget(self.prec_tabla, 1)

        # ── Barra de controles (2 filas) ──────────────────────
        ctrl_w = QWidget()
        ctrl_lay = QHBoxLayout(ctrl_w)
        ctrl_lay.setContentsMargins(0, 4, 0, 0)
        ctrl_lay.setSpacing(8)

        btn_sel = QPushButton("☑  Todos")
        btn_sel.setObjectName("btn_secundario")
        btn_sel.clicked.connect(lambda: self._prec_seleccionar(True))
        btn_desel = QPushButton("☐  Ninguno")
        btn_desel.setObjectName("btn_secundario")
        btn_desel.clicked.connect(lambda: self._prec_seleccionar(False))
        ctrl_lay.addWidget(btn_sel)
        ctrl_lay.addWidget(btn_desel)

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setStyleSheet("color:#444;")
        ctrl_lay.addWidget(sep1)

        ctrl_lay.addWidget(QLabel("% a seleccionados:"))
        self.prec_spin_pct = QDoubleSpinBox()
        self.prec_spin_pct.setRange(-99, 9999)
        self.prec_spin_pct.setDecimals(1)
        self.prec_spin_pct.setSingleStep(5)
        self.prec_spin_pct.setSuffix(" %")
        self.prec_spin_pct.setMinimumWidth(90)
        ctrl_lay.addWidget(self.prec_spin_pct)
        btn_aplicar_pct = QPushButton("Aplicar %")
        btn_aplicar_pct.setStyleSheet(
            "QPushButton{background:#1565C0;color:white;font-weight:600;"
            "border-radius:5px;padding:5px 12px;}"
            "QPushButton:hover{background:#1976D2;}")
        btn_aplicar_pct.clicked.connect(self._prec_aplicar_pct)
        ctrl_lay.addWidget(btn_aplicar_pct)
        ctrl_lay.addStretch()

        btn_confirmar = QPushButton("✅  Confirmar cambios")
        btn_confirmar.setStyleSheet(
            "QPushButton{background:#2E7D32;color:white;font-weight:700;"
            "border-radius:6px;padding:6px 18px;}"
            "QPushButton:hover{background:#388E3C;}")
        btn_confirmar.clicked.connect(self._prec_confirmar)
        ctrl_lay.addWidget(btn_confirmar)

        lay.addWidget(ctrl_w)
        self._prec_spins = {}   # {producto_id: (row, spin)}
        return w

    def _prec_cargar(self):
        """Carga todos los productos en la tabla de precios."""
        self._prec_spins = {}
        self.prec_tabla.setRowCount(0)
        productos = self.todos_productos if hasattr(self, "todos_productos") else db.obtener_todos_productos()

        self.prec_tabla.setRowCount(len(productos))
        for i, p in enumerate(productos):
            pid = p["id"]

            # Col 0: checkbox
            chk = QTableWidgetItem()
            chk.setCheckState(Qt.CheckState.Unchecked)
            chk.setData(Qt.ItemDataRole.UserRole, pid)
            chk.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.prec_tabla.setItem(i, 0, chk)

            # Col 1: nombre
            self.prec_tabla.setItem(i, 1, QTableWidgetItem(p["nombre"]))

            # Col 2: categoría
            cat = p["categoria_nombre"] if "categoria_nombre" in p.keys() else "—"
            self.prec_tabla.setItem(i, 2, QTableWidgetItem(cat or "—"))

            # Col 3: precio actual (readonly, guarda valor en UserRole)
            precio_actual = p["precio_venta"] or 0
            it_precio = QTableWidgetItem(f"$ {precio_actual:,.2f}")
            it_precio.setData(Qt.ItemDataRole.UserRole, precio_actual)
            it_precio.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.prec_tabla.setItem(i, 3, it_precio)

            # Col 4: nuevo precio (spinbox editable individualmente)
            spin = QDoubleSpinBox()
            spin.setRange(0, 9999999)
            spin.setDecimals(2)
            spin.setSingleStep(100)
            spin.setPrefix("$ ")
            spin.setValue(precio_actual)
            spin.setStyleSheet(
                "QDoubleSpinBox{background:#2C2C2C;color:#F5F5F5;"
                "border:1px solid #444;border-radius:4px;padding:2px 4px;}")
            spin.valueChanged.connect(
                lambda v, row=i, orig=precio_actual:
                    self._prec_actualizar_cambio(row, v, orig))
            self.prec_tabla.setCellWidget(i, 4, spin)

            # Col 5: cambio %
            it_cambio = QTableWidgetItem("—")
            it_cambio.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_cambio.setForeground(QColor("#666"))
            self.prec_tabla.setItem(i, 5, it_cambio)

            self.prec_tabla.setRowHeight(i, 44)
            self._prec_spins[pid] = (i, spin)

    def _prec_actualizar_cambio(self, row: int, nuevo: float, original: float):
        it = self.prec_tabla.item(row, 5)
        if not it:
            return
        if original == 0 or abs(nuevo - original) < 0.01:
            it.setText("—")
            it.setForeground(QColor("#666"))
        elif nuevo > original:
            pct = (nuevo - original) / original * 100
            it.setText(f"▲ +{pct:.1f}%")
            it.setForeground(QColor("#4CAF50"))
        else:
            pct = (original - nuevo) / original * 100
            it.setText(f"▼ -{pct:.1f}%")
            it.setForeground(QColor("#F44336"))

    def _prec_filtrar(self):
        texto  = self.prec_txt_filtro.text().lower()
        cat_id = self.prec_cmb_cat.currentData()
        for row in range(self.prec_tabla.rowCount()):
            chk_item = self.prec_tabla.item(row, 0)
            nom_item  = self.prec_tabla.item(row, 1)
            if not chk_item or not nom_item:
                continue
            pid = chk_item.data(Qt.ItemDataRole.UserRole)
            p   = next((x for x in self.todos_productos if x["id"] == pid), None)
            if not p:
                continue
            match_texto = not texto or texto in p["nombre"].lower()
            match_cat   = cat_id is None or p.get("categoria_id") == cat_id
            self.prec_tabla.setRowHidden(row, not (match_texto and match_cat))

    def _prec_seleccionar(self, seleccionar: bool):
        estado = Qt.CheckState.Checked if seleccionar else Qt.CheckState.Unchecked
        for row in range(self.prec_tabla.rowCount()):
            if not self.prec_tabla.isRowHidden(row):
                chk = self.prec_tabla.item(row, 0)
                if chk:
                    chk.setCheckState(estado)

    def _prec_aplicar_pct(self):
        pct = self.prec_spin_pct.value()
        if pct == 0:
            return
        aplicados = 0
        for row in range(self.prec_tabla.rowCount()):
            if self.prec_tabla.isRowHidden(row):
                continue
            chk = self.prec_tabla.item(row, 0)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                original = self.prec_tabla.item(row, 3).data(Qt.ItemDataRole.UserRole)
                spin = self.prec_tabla.cellWidget(row, 4)
                if spin:
                    spin.setValue(round(original * (1 + pct / 100), 2))
                    aplicados += 1
        if aplicados == 0:
            QMessageBox.information(self, "Sin selección",
                "Marcá al menos un producto (☑) antes de aplicar.")

    def _prec_confirmar(self):
        cambios = []
        for pid, (row, spin) in self._prec_spins.items():
            if self.prec_tabla.isRowHidden(row):
                continue
            nuevo    = spin.value()
            original = self.prec_tabla.item(row, 3).data(Qt.ItemDataRole.UserRole)
            if abs(nuevo - original) > 0.01:
                nombre = self.prec_tabla.item(row, 1).text()
                cambios.append((pid, nombre, original, nuevo))

        if not cambios:
            QMessageBox.information(self, "Sin cambios",
                "No hay precios modificados. Editá el 'Nuevo precio' de algún producto.")
            return

        resumen = "\n".join(
            f"• {nombre}:  $ {orig:,.2f}  →  $ {nuevo:,.2f}"
            for _, nombre, orig, nuevo in cambios[:12]
        )
        if len(cambios) > 12:
            resumen += f"\n… y {len(cambios) - 12} más"

        resp = QMessageBox.question(
            self, f"Confirmar {len(cambios)} cambio(s) de precio",
            f"{resumen}\n\n¿Guardar estos cambios?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if resp == QMessageBox.StandardButton.Yes:
            for pid, _, _, nuevo in cambios:
                db.actualizar_producto(pid, {"precio_venta": nuevo})
            self.cargar_productos()
            self._prec_cargar()
            QMessageBox.information(
                self, "✅ Listo",
                f"{len(cambios)} precio(s) actualizados correctamente.")

    def _aumentar_precio_masivo(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("📈  Actualizar precios por porcentaje")
        dlg.setMinimumWidth(400)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(12)
        lay.setContentsMargins(24, 20, 24, 20)

        lbl = QLabel("Aplicá un aumento o descuento porcentual a los precios de venta.")
        lbl.setWordWrap(True)
        lbl.setStyleSheet("color:#AAAAAA; font-size:10pt;")
        lay.addWidget(lbl)

        form = QFormLayout()
        form.setSpacing(12)

        spin_pct = QDoubleSpinBox()
        spin_pct.setRange(-99, 9999)
        spin_pct.setDecimals(1)
        spin_pct.setSingleStep(5)
        spin_pct.setSuffix(" %")
        spin_pct.setValue(0)
        spin_pct.setToolTip("Positivo = aumento, negativo = descuento")
        form.addRow("Porcentaje:", spin_pct)

        cmb_cat = QComboBox()
        cmb_cat.addItem("Todos los productos", None)
        for c in db.obtener_categorias():
            cmb_cat.addItem(c["nombre"], c["id"])
        form.addRow("Aplicar a:", cmb_cat)

        lbl_preview = QLabel("")
        lbl_preview.setStyleSheet("color:#C9A84C; font-size:10pt; font-weight:700;")
        form.addRow("", lbl_preview)

        lay.addLayout(form)

        def _actualizar_preview():
            pct = spin_pct.value()
            cat_txt = cmb_cat.currentText()
            if pct == 0:
                lbl_preview.setText("")
                return
            signo = "▲" if pct > 0 else "▼"
            color  = "#4CAF50" if pct > 0 else "#F44336"
            lbl_preview.setStyleSheet(f"color:{color}; font-size:10pt; font-weight:700;")
            lbl_preview.setText(
                f"{signo} {abs(pct):.1f}% sobre {cat_txt}  "
                f"(ej. $ 1.000 → $ {1000 * (1 + pct/100):,.0f})")

        spin_pct.valueChanged.connect(_actualizar_preview)
        cmb_cat.currentIndexChanged.connect(_actualizar_preview)

        btn_row = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(dlg.reject)
        btn_ok = QPushButton("✅  Aplicar")
        btn_ok.setStyleSheet(
            "QPushButton{background:#1565C0;color:white;font-weight:700;"
            "border-radius:6px;padding:8px 20px;}"
            "QPushButton:hover{background:#1976D2;}")
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            pct = spin_pct.value()
            if pct == 0:
                return
            cat_id = cmb_cat.currentData()
            n_prods = len([p for p in self.todos_productos
                           if cat_id is None or p["categoria_id"] == cat_id])
            resp = QMessageBox.question(
                self, "Confirmar actualización masiva",
                f"¿Aplicás {'+' if pct > 0 else ''}{pct:.1f}% de precio a "
                f"{n_prods} producto(s)?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if resp == QMessageBox.StandardButton.Yes:
                db.actualizar_precios_masivo(pct, cat_id)
                self.cargar_productos()
                QMessageBox.information(
                    self, "✅ Hecho",
                    f"Precios actualizados ({'+' if pct > 0 else ''}{pct:.1f}%).")

    def _agregar_stock_rapido(self, producto: dict):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Agregar stock: {producto['nombre']}")
        dlg.setMinimumWidth(380)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 20, 20, 20)

        lay.addWidget(QLabel(f"<b>{producto['nombre']}</b>"))
        lay.addWidget(QLabel(f"Stock actual: <b>{producto['stock_actual']}</b>"))

        linea = QFrame()
        linea.setFrameShape(QFrame.Shape.HLine)
        linea.setStyleSheet("color:#333;")
        lay.addWidget(linea)

        form = QFormLayout()
        form.setSpacing(10)

        spin = QSpinBox()
        spin.setRange(1, 99999)
        spin.setValue(1)
        form.addRow("Cantidad a agregar:", spin)

        txt_motivo = QLineEdit("Ingreso de mercadería")
        form.addRow("Motivo:", txt_motivo)

        # ── Cosecha (año de producción, opcional) ─────────────
        spin_cosecha = QSpinBox()
        spin_cosecha.setRange(0, 2100)
        spin_cosecha.setSpecialValueText("— Sin cosecha")
        spin_cosecha.setValue(0)
        spin_cosecha.setToolTip("Año de producción del vino. Dejá en 0 si no aplica.")
        form.addRow("🍷 Cosecha (año):", spin_cosecha)

        # ── Fecha de vencimiento (opcional) ───────────────────
        chk_venc = QCheckBox("Tiene fecha de vencimiento")
        form.addRow("", chk_venc)

        date_venc = QDateEdit()
        date_venc.setCalendarPopup(True)
        date_venc.setDisplayFormat("dd/MM/yyyy")
        date_venc.setDate(QDate.currentDate().addMonths(6))
        date_venc.setEnabled(False)
        chk_venc.stateChanged.connect(lambda s: date_venc.setEnabled(bool(s)))
        form.addRow("📅 Fecha de vencimiento:", date_venc)

        lay.addLayout(form)

        btn_row = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setObjectName("btn_secundario")
        btn_cancel.clicked.connect(dlg.reject)
        btn_ok = QPushButton("✅  Agregar stock")
        btn_ok.setStyleSheet(
            "QPushButton{background:#2E7D32;color:white;font-weight:700;"
            "border-radius:6px;padding:8px 20px;}"
            "QPushButton:hover{background:#388E3C;}"
        )
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            cantidad  = spin.value()
            motivo    = txt_motivo.text().strip() or "Ingreso de mercadería"
            cosecha   = spin_cosecha.value() if spin_cosecha.value() > 0 else None
            fecha_v   = date_venc.date().toString("yyyy-MM-dd") if chk_venc.isChecked() else None
            db.agregar_stock(producto["id"], cantidad, motivo)
            if cosecha or fecha_v:
                db.crear_lote(producto["id"], cantidad,
                              cosecha=cosecha, fecha_vencimiento=fecha_v, motivo=motivo)
            self.cargar_productos()

    def _eliminar_producto(self, producto: dict):
        resp = QMessageBox.question(
            self, "Eliminar producto",
            f"¿Eliminás '{producto['nombre']}'?\n"
            "El historial de ventas se conserva.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if resp == QMessageBox.StandardButton.Yes:
            db.eliminar_producto(producto["id"])
            self.cargar_productos()

    def _cargar_stock(self):
        dlg = DialogoCargaStock(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.cargar_productos()

    def _ajustar_stock_producto(self, producto: dict):
        dlg = DialogoAjusteStock(self, producto)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            nuevo_stock, motivo = dlg.get_values()
            if nuevo_stock != producto["stock_actual"]:
                db.ajustar_stock(producto["id"], nuevo_stock, motivo)
                self.cargar_productos()
            elif dlg.txt_notas.text().strip():
                QMessageBox.information(self, "Sin cambio",
                    "El stock no fue modificado (el nuevo valor es igual al actual).")

    def _ver_lotes(self, producto: dict):
        dlg = DialogoLotesProducto(self, producto)
        dlg.exec()

    def _cargar_vencimientos(self):
        from datetime import date
        hoy  = date.today()
        dias = self.spin_dias_venc.value()

        vencidos  = db.lotes_vencidos()
        proximos  = db.lotes_por_vencer(dias)
        # Desduplicar (vencidos + próximos no se solapan por definición, pero por seguridad)
        vistos, todos = set(), []
        for l in list(vencidos) + list(proximos):
            if l["id"] not in vistos:
                vistos.add(l["id"])
                todos.append(l)

        self.tabla_vencimientos.setRowCount(len(todos))
        for i, l in enumerate(todos):
            self.tabla_vencimientos.setItem(i, 0, QTableWidgetItem(l["producto_nombre"]))
            cosecha = str(l["cosecha"]) if l["cosecha"] else "—"
            self.tabla_vencimientos.setItem(i, 1, QTableWidgetItem(cosecha))

            cant_item = QTableWidgetItem(str(l["cantidad"]))
            cant_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla_vencimientos.setItem(i, 2, cant_item)
            self.tabla_vencimientos.setItem(i, 3, QTableWidgetItem(l["fecha_vencimiento"]))

            dias_rest = (date.fromisoformat(l["fecha_vencimiento"]) - hoy).days
            if dias_rest < 0:
                dias_txt = f"VENCIDO hace {-dias_rest}d"
                color    = QColor("#9E9E9E")
            elif dias_rest == 0:
                dias_txt = "Vence HOY"
                color    = QColor("#EF5350")
            elif dias_rest <= 7:
                dias_txt = f"{dias_rest} días"
                color    = QColor("#EF5350")
            elif dias_rest <= 30:
                dias_txt = f"{dias_rest} días"
                color    = QColor("#FF9800")
            else:
                dias_txt = f"{dias_rest} días"
                color    = QColor("#66BB6A")

            dias_item = QTableWidgetItem(dias_txt)
            dias_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            dias_item.setForeground(color)
            self.tabla_vencimientos.setItem(i, 4, dias_item)
            self.tabla_vencimientos.setItem(i, 5, QTableWidgetItem(l["motivo"] or ""))
            self.tabla_vencimientos.setRowHeight(i, 36)

        if not todos:
            self.tabla_vencimientos.setRowCount(1)
            item = QTableWidgetItem(f"✅  Sin vencimientos en los próximos {dias} días ni vencidos")
            item.setForeground(QColor("#4CAF50"))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla_vencimientos.setItem(0, 0, item)
            self.tabla_vencimientos.setSpan(0, 0, 1, 6)
